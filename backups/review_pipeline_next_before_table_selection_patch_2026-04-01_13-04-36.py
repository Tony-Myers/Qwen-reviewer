#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

import pdfplumber
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from mlx_lm import load, generate
from mlx_lm.sample_utils import make_sampler


MODEL_NAME = "mlx-community/Qwen3.5-35B-A3B-4bit"
OUTPUT_DOMAIN = "general"

MAX_SECTION_CHARS = 7000
SECTION_MAX_TOKENS = 900
FILE_SYNTHESIS_MAX_TOKENS = 1200
SYNTHESIS_MAX_TOKENS = 1600
VALIDATION_MAX_TOKENS = 1800

TEMPERATURE = 0.2
TOP_P = 0.8
TOP_K = 20

MAX_TABLE_ROWS = 80
MAX_TABLE_COLS = 14
MAX_TABLE_CHARS_PER_CELL = 120
MAX_APPENDIX_CHUNK_PREVIEW = 1800


SYSTEM_STYLE = """You are a rigorous academic peer review assistant.

Core rules:
- Work strictly from the provided material.
- Distinguish clearly between:
  1. Directly supported by the text
  2. Reasonable inference
  3. Speculation / missing information
- Do not invent sample sizes, analyses, results, measures, design features, claims, contradictions, artefacts, or errors.
- Prioritise internal validity, measurement quality, causal assumptions, statistical reasoning, modelling choices, transparency, and reporting.
- Be critical but fair.
- Use concise markdown headings and bullet points.
- Do NOT use LaTeX or dollar-sign math notation.
- If tabular material is present, treat it as available evidence unless clearly unreadable.
- If a table is imperfectly extracted, say "partly extracted" or "some entries may be ambiguous", not "missing", unless it is genuinely absent.

Generality rule:
- Apply the same standards across frequentist, Bayesian, predictive, and causal analyses.
- Do not assume a method-specific problem unless the extracted material supports it.

Severity ladder:
- Prefer "could be clarified" over "is wrong" unless the text directly supports the stronger claim.
- Prefer "reported briefly" or "not fully detailed" over "unreported" when something is mentioned but not elaborated.
- Prefer "interpretation may need clarification" over "data entry artefact" unless there is direct evidence of corruption.
- Prefer "could not verify from the extracted material" over "missing" when extraction is partial.
"""

COMMON_DIAGNOSTIC_ALIASES = """Examples of reported diagnostics and checks across paradigms:
- frequentist/model-based: VIF, residual plots, heteroscedasticity checks, AIC, BIC, confidence intervals, sensitivity analyses, multiple imputation, cross-validation
- Bayesian: priors, prior predictive checks, posterior predictive checks, R-hat, effective sample size, divergences, LOO, WAIC, Bayes factors, prior sensitivity
- prediction/ML: calibration, discrimination, ROC/AUC, Brier score, internal validation, external validation, optimism correction
- causal/design: DAGs, assumptions, preregistration, protocol deviations, missing-data mechanisms
"""


@dataclass
class DocChunk:
    source_name: str
    chunk_id: int
    text: str


SUPERSCRIPT_MAP = str.maketrans({
    "0": "⁰", "1": "¹", "2": "²", "3": "³", "4": "⁴",
    "5": "⁵", "6": "⁶", "7": "⁷", "8": "⁸", "9": "⁹",
    "+": "⁺", "-": "⁻", "=": "⁼", "(": "⁽", ")": "⁾",
    "n": "ⁿ", "i": "ⁱ",
})

SUBSCRIPT_MAP = str.maketrans({
    "0": "₀", "1": "₁", "2": "₂", "3": "₃", "4": "₄",
    "5": "₅", "6": "₆", "7": "₇", "8": "₈", "9": "₉",
    "+": "₊", "-": "₋", "=": "₌", "(": "₍", ")": "₎",
    "a": "ₐ", "e": "ₑ", "i": "ᵢ", "j": "ⱼ", "o": "ₒ",
    "r": "ᵣ", "u": "ᵤ", "v": "ᵥ", "x": "ₓ",
})


def _to_superscript(s: str) -> str:
    return s.translate(SUPERSCRIPT_MAP)


def _to_subscript(s: str) -> str:
    return s.translate(SUBSCRIPT_MAP)


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("￾", "")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clean_extracted_evidence_text(text: str, domain: str = "general") -> str:
    text = clean_text(text)

    replacements = {
        "VO 2max": "VO2max",
        "VO 2peak": "VO2peak",
        "VO ₂max": "VO2max",
        "VO ₂peak": "VO2peak",
        "ln (": "ln(",
        "Ln (": "ln(",
        "Age 2": "Age2",
        "age 2": "age2",
        "L. min−1": "L·min−1",
        "mL. kg−1.min−1": "mL·kg−1·min−1",
        "kg −1": "kg−1",
        "min −1": "min−1",
        "body fat %": "body fat%",
        "BF %": "BF%",
        "Cyc le": "Cycle",
        "algor ithms": "algorithms",
        "method- ology": "methodology",
        "allome- tric": "allometric",
        "physiolog- ical": "physiological",
        "incor - porating": "incorporating",
        "log- transformed": "log-transformed",
        "out- of- sample": "out-of-sample",
        "all- cause": "all-cause",
        "fat- free": "fat-free",
        "whole- body": "whole-body",
        "body- composition": "body-composition",
        "X- ray": "X-ray",
        "meta- analysis": "meta-analysis",
        "cross- validation": "cross-validation",
        "peer- reviewed": "peer-reviewed",
        "non- linear": "non-linear",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"([A-Za-z])-\s+([A-Za-z])", r"\1\2", text)
    text = re.sub(r"ln\(mass\(\s*g\s*\)\)", "ln(mass)", text, flags=re.IGNORECASE)
    text = re.sub(r"ln\(mass\s*\(\s*g\s*\)\)", "ln(mass)", text, flags=re.IGNORECASE)
    text = re.sub(r"\bVO\s*\(?\s*2max\s*\)?", "VO2max", text)

    if domain == "exercise_physiology":
        text = re.sub(r"\bln\(masskg\)\b", "ln(mass kg)", text, flags=re.IGNORECASE)
        text = re.sub(r"\bln\(fffkg\)\b", "ln(FFM kg)", text, flags=re.IGNORECASE)
        text = re.sub(r"\blnfvc\b", "ln(FVC)", text, flags=re.IGNORECASE)
        text = re.sub(r"\blnwc\b", "ln(WC)", text, flags=re.IGNORECASE)

    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def truncate_cell(value, max_chars: int = MAX_TABLE_CHARS_PER_CELL) -> str:
    if value is None:
        return ""
    s = str(value).strip().replace("\n", " ").replace("\t", " ")
    s = re.sub(r"\s{2,}", " ", s)
    if len(s) > max_chars:
        s = s[: max_chars - 3] + "..."
    return s


def table_to_block(table: List[List[str]], table_num: int, page_num: int) -> str:
    if not table:
        return ""

    trimmed_rows = table[:MAX_TABLE_ROWS]
    lines = [
        "[TABLE_START]",
        f"Page: {page_num}",
        f"Table number on page: {table_num}",
    ]
    for row in trimmed_rows:
        row = row[:MAX_TABLE_COLS]
        cleaned = [truncate_cell(cell) for cell in row]
        lines.append("\t".join(cleaned))
    lines.append("[TABLE_END]")
    return "\n".join(lines).strip()


def extract_tables_from_pdf(path: Path) -> List[Tuple[int, str]]:
    table_blocks: List[Tuple[int, str]] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                extracted_tables = []
                try:
                    raw_tables = page.extract_tables()
                except Exception:
                    raw_tables = []
                table_counter = 0
                for tbl in raw_tables or []:
                    if not tbl:
                        continue
                    tbl = [
                        ["" if cell is None else str(cell) for cell in row]
                        for row in tbl
                        if row and any((cell is not None and str(cell).strip()) for cell in row)
                    ]
                    if len(tbl) < 2:
                        continue
                    table_counter += 1
                    block = table_to_block(tbl, table_counter, page_index)
                    if block:
                        extracted_tables.append(block)
                if extracted_tables:
                    table_blocks.append((page_index, "\n\n".join(extracted_tables)))
    except Exception:
        pass
    return table_blocks


def read_pdf(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    reader = PdfReader(str(path))
    table_blocks = extract_tables_from_pdf(path)
    table_map = {page_num: txt for page_num, txt in table_blocks}
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as e:
            text = f"[Could not extract text from page {i}: {e}]"
        cleaned_page_text = clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN)
        page_parts = [f"\n\n[Page {i}]\n{cleaned_page_text}"]
        if i in table_map:
            page_parts.append("\n[Extracted tables]\n" + table_map[i])
        pages.append("\n".join(part for part in page_parts if part.strip()))
    return "\n".join(pages).strip(), table_blocks


def read_txt(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    return clean_extracted_evidence_text(
        path.read_text(encoding="utf-8", errors="ignore"),
        domain=OUTPUT_DOMAIN,
    ), []


def read_docx(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs)
    return clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN), []


def read_csv_file(path: Path, max_rows: int = 120) -> Tuple[str, List[Tuple[int, str]]]:
    rows = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append("\t".join("" if x is None else str(x) for x in row))
    text = f"[CSV preview: first {min(len(rows), max_rows)} rows]\n" + "\n".join(rows)
    return clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN), []


def read_xlsx(path: Path, max_rows_per_sheet: int = 80) -> Tuple[str, List[Tuple[int, str]]]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"\n[Sheet: {ws.title}]")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                break
            out.append("\t".join("" if v is None else str(v) for v in row))
    return clean_extracted_evidence_text("\n".join(out), domain=OUTPUT_DOMAIN), []


def load_document(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf(path)
    if suffix in {".txt", ".md"}:
        return read_txt(path)
    if suffix == ".docx":
        return read_docx(path)
    if suffix == ".csv":
        return read_csv_file(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError(f"Unsupported file type: {path.name}")


def split_text(text: str, max_chars: int = MAX_SECTION_CHARS) -> List[str]:
    text = clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN)
    if not text:
        return ["[No extractable text found]"]
    if len(text) <= max_chars:
        return [text]
    paras = text.split("\n\n")
    chunks: List[str] = []
    current: List[str] = []
    current_len = 0
    for para in paras:
        para = para.strip()
        if not para:
            continue
        add_len = len(para) + 2
        if current and current_len + add_len > max_chars:
            chunks.append("\n\n".join(current).strip())
            current = [para]
            current_len = len(para)
        else:
            current.append(para)
            current_len += add_len
    if current:
        chunks.append("\n\n".join(current).strip())
    return chunks if chunks else ["[No extractable text found]"]


def collect_input_paths(paths: Iterable[str]) -> List[Path]:
    supported = {".pdf", ".txt", ".md", ".docx", ".csv", ".xlsx", ".xlsm"}
    results: List[Path] = []
    for raw in paths:
        p = Path(raw).expanduser()
        if p.is_dir():
            for child in sorted(p.rglob("*")):
                if child.is_file() and child.suffix.lower() in supported:
                    results.append(child)
        elif p.is_file() and p.suffix.lower() in supported:
            results.append(p)
    return results


def apply_qwen_chat_template(tokenizer, user_text: str) -> str:
    messages = [{"role": "user", "content": user_text}]
    return tokenizer.apply_chat_template(
        messages,
        tokenize=False,
        add_generation_prompt=True,
        enable_thinking=False,
    )


def make_default_sampler():
    return make_sampler(TEMPERATURE, top_p=TOP_P, top_k=TOP_K)


def basic_cleanup(text: str) -> str:
    text = unicodedata.normalize("NFKC", text)
    replacements = {
        "\x00": " ",
        "￾": "",
        "$": "",
        "\\%": "%",
        "\\_": "_",
        "\\&": "&",
        "\\#": "#",
        "\\$": "$",
        "\\(": "",
        "\\)": "",
        "\\[": "",
        "\\]": "",
        "\\cdot": "·",
        "\\times": "×",
        "\\approx": "≈",
        "\\geq": "≥",
        "\\leq": "≤",
        "\\neq": "≠",
        "\\pm": "±",
        "\\to": "→",
        "−": "−",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = text.replace("```markdown", "").replace("```text", "").replace("```", "")
    text = re.sub(r"\{([^{}]{1,50})\}", r"\1", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def scientific_cleanup(text: str) -> str:
    replacements = {
        "\\alpha": "α",
        "\\beta": "β",
        "\\gamma": "γ",
        "\\delta": "δ",
        "\\Delta": "Δ",
        "\\mu": "μ",
        "\\sigma": "σ",
        "\\lambda": "λ",
        "\\theta": "θ",
        "\\rho": "ρ",
        "\\tau": "τ",
        "\\chi": "χ",
        "\\phi": "φ",
        "\\psi": "ψ",
        "\\omega": "ω",
        "R^2": "R²",
        "r^2": "r²",
        "p-value": "p value",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"(\d+)\s*\\?%\s*CI", r"\1% CI", text)
    text = re.sub(r"\bCI\s*:", "CI:", text)
    text = re.sub(r"\bp\s*=\s*", "p = ", text)
    text = re.sub(r"\bp\s*<\s*", "p < ", text)
    text = re.sub(r"\bp\s*>\s*", "p > ", text)
    text = re.sub(r"\bp\s*≤\s*", "p ≤ ", text)
    text = re.sub(r"\bp\s*≥\s*", "p ≥ ", text)
    text = re.sub(
        r"([A-Za-z%°·/\)])\^\{([0-9+\-=\(\)ni]+)\}",
        lambda m: m.group(1) + _to_superscript(m.group(2)),
        text,
    )
    text = re.sub(
        r"([A-Za-z%°·/\)])\^([0-9+\-=\(\)ni]{1,4})\b",
        lambda m: m.group(1) + _to_superscript(m.group(2)),
        text,
    )
    text = re.sub(
        r"([A-Za-z])_\{([0-9a-zA-Z+\-\(\)=]{1,6})\}",
        lambda m: m.group(1) + _to_subscript(m.group(2).lower()),
        text,
    )
    text = re.sub(
        r"([A-Za-z])_([0-9a-zA-Z]{1,4})\b",
        lambda m: m.group(1) + _to_subscript(m.group(2).lower()),
        text,
    )
    text = re.sub(r"\b([mk]?g)\s*−\s*1\b", r"\1⁻¹", text)
    text = re.sub(r"\b(min|s|m|L)\s*−\s*1\b", r"\1⁻¹", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def exercise_physiology_cleanup(text: str) -> str:
    replacements = {
        "VO₂max": "VO2max",
        "VO₂peak": "VO2peak",
        "VO₂": "VO2",
        "VCO₂": "VCO2",
        "BF\\%": "BF%",
        "BF%": "BF%",
        "Ln(": "ln(",
        "Ln ": "ln ",
        "Age^2": "Age²",
        "age^2": "age²",
        "FFMkg": "FFM (kg)",
        "Lnmasskg": "ln(mass kg)",
        "LnFFFkg": "ln(FFM kg)",
        "LnFVC": "ln(FVC)",
        "LnWC": "ln(WC)",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"VO2max\s*\(\s*L\s*[·\.]\s*min⁻¹\s*\)", "VO2max (L·min⁻¹)", text)
    text = re.sub(r"VO2max\s*\(\s*mL\s*[·\.]\s*kg⁻¹\s*[·\.]\s*min⁻¹\s*\)", "VO2max (mL·kg⁻¹·min⁻¹)", text)
    return text.strip()


def clean_model_output(text: str, domain: str = "general") -> str:
    text = basic_cleanup(text)
    text = scientific_cleanup(text)
    if domain == "exercise_physiology":
        text = exercise_physiology_cleanup(text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def review_chunk(model, tokenizer, chunk: DocChunk) -> str:
    user_text = f"""
{SYSTEM_STYLE}

{COMMON_DIAGNOSTIC_ALIASES}

Review this manuscript chunk.

Source: {chunk.source_name}
Chunk: {chunk.chunk_id}

Return markdown under exactly these headings:
## Supported points
## Reasonable inferences
## Missing or unclear
## Statistical / methodological concerns
## Table-specific notes
## Questions raised by this chunk

Instructions:
- Do not infer that an analysis is Bayesian merely because the text mentions MCMC, Gibbs sampling, or Markov Chain Monte Carlo.
- Treat inferential framework and computational method as separate issues.
- Only apply Bayesian-reporting expectations when the text explicitly indicates Bayesian analysis, such as priors, posteriors, credible intervals, Bayes factors, posterior predictive checks, or similar language.
- If MCMC is mentioned without clear Bayesian framing, describe the framework as unclear or computationally MCMC-based, not automatically Bayesian.
- Where tables are present, use them as evidence.
- If a table appears partially extracted, say so explicitly.
- Do not claim a table, coefficient, equation, or diagnostic is missing if it is present in the chunk.
- Do not use LaTeX or dollar-sign notation.
- Be concise and avoid repetition.
- Do not ask for information already clearly reported in the chunk.
- Do not convert obvious PDF/OCR extraction artefacts into substantive methodological concerns.
- Do not describe a parameterisation choice or scaling convention as an error unless the text directly contradicts itself.
- Do not infer data entry artefacts from a small coefficient with a large test statistic alone.
- If an equation or model form is shown in the chunk, do not say the functional form is absent or unclear.
- Distinguish between "not reported", "reported briefly", and "reported but not fully interpretable from this chunk".
- If the text refers to an equation, figure, table, appendix item, or model specification that is not clearly recoverable from extraction, do not call it missing.
- Instead say that it appears to be present in the PDF but could not be evaluated confidently from the extracted material.
- Where appropriate, direct the human reviewer to inspect the original PDF.

Material:
\"\"\"
{chunk.text}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=SECTION_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out, domain=OUTPUT_DOMAIN)


def synthesize_file_review(model, tokenizer, file_name: str, combined_chunk_review: str) -> str:
    user_text = f"""
{SYSTEM_STYLE}

{COMMON_DIAGNOSTIC_ALIASES}

Create a concise file-level review summary from the chunk notes below.

File: {file_name}

Required headings:
# File synopsis
# Strongly supported strengths
# Strongly supported concerns
# Reporting limits or ambiguities
# Questions raised by this file

Rules:
- Distinguish between inferential framework and computational method.
- Do not classify a study as Bayesian solely because it uses MCMC or MLwiN.
- Only criticise missing priors, posterior diagnostics, or Bayes factors when the manuscript clearly uses a Bayesian framework.
- Base the summary only on the chunk notes.
- Maximum 4 bullets under any heading.
- Prefer directly supported points over generic reviewer concerns.
- Do not say a table, coefficient, equation, or diagnostic is missing if the chunk notes contain it.
- If something is present but brief, say "reported briefly" or "not fully detailed", not "unreported".
- Do not use LaTeX or dollar-sign notation.
- Do not convert obvious PDF/OCR extraction artefacts into substantive concerns.
- Do not call a reporting choice a contradiction unless two directly conflicting statements are present.
- Do not infer data entry artefacts from small coefficients, exponent notation, or scaling choices alone.
- Finish cleanly after the final heading.
- Distinguish between "not present in the manuscript" and "not recoverably extracted".
- If an item is referenced in the manuscript text but not clearly accessible from extraction, describe it as "present but not fully accessible in the extracted material".
- Treat extraction failure as a review limitation, not automatically as a manuscript fault.

Chunk notes:
\"\"\"
{combined_chunk_review}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=FILE_SYNTHESIS_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out, domain=OUTPUT_DOMAIN)


def synthesize_report(model, tokenizer, file_summaries: List[Tuple[str, str]]) -> str:
    joined = []
    for name, summary in file_summaries:
        joined.append(f"# File: {name}\n{summary}")
    joined_text = "\n\n".join(joined)

    user_text = f"""
{SYSTEM_STYLE}

{COMMON_DIAGNOSTIC_ALIASES}

Create one integrated peer-review report from the file summaries below.

Required headings:
# Overall synopsis
# Major strengths
# Major concerns
# Statistical and methodological issues
# Table and supplementary material issues
# Questions for the authors
# Suggested comments to editor
# Confidence and limits of this review

Rules:
- Do not infer Bayesian analysis from MCMC alone.
- Only include Bayesian-specific criticisms if the file summaries clearly indicate Bayesian modelling through priors, posteriors, credible intervals, Bayes factors, posterior predictive checks, or equivalent terms.
- Base the report only on the supplied file summaries.
- State clearly when a point is directly supported versus inferred.
- Do not overclaim.
- Maximum 4 bullets under any heading.
- Prefer concise bullets over long paragraphs.
- Do not repeat the same point across multiple headings unless necessary.
- Do not use LaTeX or dollar-sign notation.
- Do not claim that a table, coefficient, equation, or diagnostic is missing if it is present in the file summaries.
- Do not ask for a statistic as unreported if it is reported in the file summaries.
- If reporting is present but limited, say "reported briefly" or "reported without full detail".
- Do not convert obvious PDF/OCR extraction artefacts into substantive concerns.
- Do not call a reporting choice a contradiction unless two directly conflicting statements are present.
- Do not infer data entry artefacts from small coefficients, exponent notation, or scaling choices alone.
- If the model form or equation appears in the summaries, do not say the functional form is absent or unclear.
- Finish cleanly after the final heading.
- If a criticism depends on the absence of an equation, figure, table, appendix item, or model specification, include it only when the evidence summaries genuinely indicate absence.
- If the manuscript text points to such an item but extraction is inadequate, say it appears to be present in the PDF but could not be fully evaluated from the extracted material.
- Treat extraction limitations as limitations of this review unless the evidence supports a stronger claim.

File summaries:
\"\"\"
{joined_text}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=SYNTHESIS_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out, domain=OUTPUT_DOMAIN)


def validate_report_against_evidence(model, tokenizer, report_text: str, file_summaries: List[Tuple[str, str]]) -> str:
    joined = []
    for name, summary in file_summaries:
        joined.append(f"# File: {name}\n{summary}")
    evidence_text = "\n\n".join(joined)

    user_text = f"""You are checking a peer-review report for factual consistency and proportionality against extracted evidence summaries.

Task:
Revise the report only where it incorrectly claims that information is missing, absent, contradictory, erroneous, or unreported.

General rule:
- Do not criticise the absence of something that is actually reported.
- Only criticise adequacy, transparency, or interpretation when the evidence supports that criticism.

{COMMON_DIAGNOSTIC_ALIASES}

Important rules:
- If the evidence mentions MCMC, Gibbs sampling, MLwiN, or similar computation methods without explicit priors/posteriors/Bayes factors/credible intervals, do not rewrite the review as if the study were definitely Bayesian.
- Remove Bayesian-specific criticisms unless the evidence clearly supports a Bayesian inferential framework.
- Treat both extracted table blocks and plain chunk text containing tabular rows as evidence.
- If tables, coefficients, standard errors, confidence intervals, fit statistics, equations, or diagnostics are present in the evidence summaries, do not describe them as missing, absent, or unverifiable.
- If something is present but concise, revise wording to "reported briefly", "not fully detailed", or "could not be fully evaluated from the extracted material".
- If a model form appears in the evidence summaries, do not say the functional form is undefined.
- Do not convert obvious PDF/OCR extraction artefacts into substantive methodological concerns.
- Do not infer data entry artefacts from small coefficients, exponent notation, or scaling choices alone.
- Do not describe table numbering as inconsistent unless the evidence summaries themselves contain conflicting labels.
- Do not treat notation or terminology conventions as a major concern unless they create genuine inferential ambiguity.
- Replace overly strong wording with proportionate wording when needed.
- Preserve valid critical points.
- Do not introduce new criticisms.
- Do not use LaTeX or dollar-sign notation.
- Return the revised report only.
- If the evidence summaries refer to Eq., Figure, Table, Appendix, Supplementary Table, or similar labels, do not say the item is missing unless the evidence explicitly indicates absence.
- Revise such claims to wording like "the item appears to be present in the PDF but could not be fully evaluated from the extracted material".
- Treat extraction failure as a review limitation rather than evidence that the manuscript omitted the item.
- Where the manuscript clearly points to an equation, table, figure, or appendix item that was not recoverably extracted, treat this as a limitation of the extracted material rather than evidence that the item is absent.

Report:
\"\"\"
{report_text}
\"\"\"

Evidence summaries:
\"\"\"
{evidence_text}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=VALIDATION_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out, domain=OUTPUT_DOMAIN)


def write_report(output_dir: Path, input_paths: List[Path], report_text: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    stem = input_paths[0].stem if len(input_paths) == 1 else "multi_file_review"
    report_path = output_dir / f"{stem}_review_{stamp}.md"
    sources = "\n".join(f"- {p}" for p in input_paths)
    header = f"""# Local peer-review report

Generated: {datetime.now().isoformat(timespec="seconds")}

Model: {MODEL_NAME}

Input files:
{sources}

---

"""
    report_path.write_text(header + report_text, encoding="utf-8")
    return report_path


def write_evidence_appendix(
    output_dir: Path,
    input_paths: List[Path],
    per_file_tables: Dict[str, List[Tuple[int, str]]],
    per_file_chunks: Dict[str, List[str]],
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    stem = input_paths[0].stem if len(input_paths) == 1 else "multi_file_review"
    appendix_path = output_dir / f"{stem}_evidence_appendix_{stamp}.md"

    lines: List[str] = []
    lines.append("# Evidence appendix")
    lines.append("")
    lines.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"Model: {MODEL_NAME}")
    lines.append("")
    lines.append("## Input files")
    for p in input_paths:
        lines.append(f"- {p}")
    lines.append("")

    for p in input_paths:
        name = p.name
        lines.append(f"# File: {name}")
        lines.append("")

        tables = per_file_tables.get(name, [])
        if tables:
            lines.append("## Extracted tables")
            lines.append("")
            for page_num, block in tables:
                lines.append(f"### Page {page_num}")
                lines.append("")
                lines.append("```text")
                lines.append(block)
                lines.append("```")
                lines.append("")
        else:
            lines.append("## Extracted tables")
            lines.append("")
            lines.append("No extracted table blocks recorded.")
            lines.append("")

        chunks = per_file_chunks.get(name, [])
        if chunks:
            lines.append("## Chunk previews")
            lines.append("")
            for i, chunk_text in enumerate(chunks, start=1):
                preview = chunk_text[:MAX_APPENDIX_CHUNK_PREVIEW]
                if len(chunk_text) > MAX_APPENDIX_CHUNK_PREVIEW:
                    preview += "\n...[truncated preview]..."
                lines.append(f"### Chunk {i}")
                lines.append("")
                lines.append("```text")
                lines.append(preview)
                lines.append("```")
                lines.append("")
        else:
            lines.append("## Chunk previews")
            lines.append("")
            lines.append("No chunk previews available.")
            lines.append("")

    appendix_path.write_text("\n".join(lines), encoding="utf-8")
    return appendix_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Local manuscript review pipeline using MLX + Qwen3.5")
    parser.add_argument("inputs", nargs="+", help="One or more files or folders")
    parser.add_argument(
        "--output-dir",
        default=str(Path.home() / "local-llm/qwen35-review/reports"),
        help="Directory for markdown reports",
    )
    args = parser.parse_args()

    input_paths = collect_input_paths(args.inputs)
    if not input_paths:
        print("No supported input files found.", file=sys.stderr)
        return 1

    print("Loading model...")
    model, tokenizer = load(MODEL_NAME)

    file_summaries: List[Tuple[str, str]] = []
    per_file_tables: Dict[str, List[Tuple[int, str]]] = {}
    per_file_chunks: Dict[str, List[str]] = {}

    for path in input_paths:
        print(f"Reading: {path}")
        try:
            text, table_blocks = load_document(path)
        except Exception as e:
            file_summaries.append((path.name, f"Failed to read file: {e}"))
            per_file_tables[path.name] = []
            per_file_chunks[path.name] = []
            continue

        per_file_tables[path.name] = table_blocks
        chunks = split_text(text)
        per_file_chunks[path.name] = chunks

        if not chunks:
            file_summaries.append((path.name, "No usable text extracted from this file."))
            continue

        chunk_outputs = []
        for i, chunk_text in enumerate(chunks, start=1):
            print(f"Reviewing {path.name} chunk {i}/{len(chunks)}")
            chunk = DocChunk(source_name=path.name, chunk_id=i, text=chunk_text)
            try:
                reviewed = review_chunk(model, tokenizer, chunk)
            except Exception as e:
                reviewed = f"Chunk review failed: {e}"
            chunk_outputs.append(reviewed)

        combined = "\n\n".join(f"### Chunk {i}\n{txt}" for i, txt in enumerate(chunk_outputs, start=1))
        print(f"Synthesising file-level review for {path.name}...")
        try:
            file_summary = synthesize_file_review(model, tokenizer, path.name, combined)
        except Exception as e:
            file_summary = f"# File synopsis\nFile-level synthesis failed: {e}"
        file_summaries.append((path.name, file_summary))

    print("Synthesising final report...")
    try:
        final_report = synthesize_report(model, tokenizer, file_summaries)
        print("Validating final report against evidence...")
        final_report = validate_report_against_evidence(model, tokenizer, final_report, file_summaries)
    except Exception as e:
        final_report = (
            "# Overall synopsis\n"
            f"Synthesis failed: {e}\n\n"
            "# File-level notes\n\n"
            + "\n\n".join(f"## {name}\n{summary}" for name, summary in file_summaries)
        )

    output_dir = Path(args.output_dir)
    report_path = write_report(output_dir, input_paths, final_report)
    appendix_path = write_evidence_appendix(output_dir, input_paths, per_file_tables, per_file_chunks)

    print(f"Saved report: {report_path}")
    print(f"Saved evidence appendix: {appendix_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
