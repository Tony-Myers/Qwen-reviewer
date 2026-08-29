#!/usr/bin/env python3
"""
Local peer-review pipeline using MLX + Qwen.

Architecture:
  1. Document parsing  (PDF/docx/csv/xlsx/txt/md, optional Marker)
  2. Evidence structuring  (typed blocks: TABLE, MODEL, DESIGN, RESULT, ...)
  3. Method classification  (rule-based framework detection)
  4. Method-aware critique  (framework-specific prompt expectations)
  5. Synthesis + validation  (LLM synthesis, programmatic + LLM post-checks)
  6. Output  (review report + evidence appendix with typed blocks)
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum, auto
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pdfplumber
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from mlx_lm import load, generate
from mlx_lm.sample_utils import make_sampler


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MODEL_NAME = "mlx-community/Qwen3.5-35B-A3B-4bit"
OUTPUT_DOMAIN = "general"

MAX_SECTION_CHARS = 7000
SECTION_MAX_TOKENS = 900
FILE_SYNTHESIS_MAX_TOKENS = 1200
SYNTHESIS_MAX_TOKENS = 1600
VALIDATION_MAX_TOKENS = 1800

TEMPERATURE = 0.2
TOP_P = 0.8
TOP_K = 20

MAX_TABLE_ROWS = 80
MAX_TABLE_COLS = 14
MAX_TABLE_CHARS_PER_CELL = 120
MAX_APPENDIX_CHUNK_PREVIEW = 1800


# ---------------------------------------------------------------------------
# Phase 2: Evidence block types
# ---------------------------------------------------------------------------

class BlockType(Enum):
    TABLE = auto()
    MODEL = auto()
    DESIGN = auto()
    RESULT = auto()
    FIGURE_CAPTION = auto()
    NARRATIVE = auto()


@dataclass
class EvidenceBlock:
    """A typed unit of extracted evidence with source tracking."""
    block_type: BlockType
    text: str
    source_name: str
    page: Optional[int] = None
    label: Optional[str] = None
    confidence: str = "high"  # high | medium | low


@dataclass
class EvidenceManifest:
    """Summary of all structured evidence available for a single file."""
    source_name: str
    blocks: List[EvidenceBlock] = field(default_factory=list)
    method_class: Optional["MethodClass"] = None
    table_labels: List[str] = field(default_factory=list)
    has_equations: bool = False
    has_model_spec: bool = False
    has_effect_sizes: bool = False
    has_confidence_intervals: bool = False
    has_p_values: bool = False
    has_sample_description: bool = False
    has_randomisation: bool = False
    has_standard_errors: bool = False
    has_variance_components: bool = False
    has_model_fit_stats: bool = False
    n_tables: int = 0
    n_figures_mentioned: int = 0

    def summary_text(self) -> str:
        """Human-readable manifest for injection into review prompts."""
        lines = [f"Evidence manifest for: {self.source_name}"]
        lines.append(f"  Method classification: {self.method_class.value if self.method_class else 'unclassified'}")
        lines.append(f"  Tables extracted: {self.n_tables} ({', '.join(self.table_labels) if self.table_labels else 'none labelled'})")
        lines.append(f"  Figures mentioned in text: {self.n_figures_mentioned}")
        lines.append(f"  Model specification present: {self.has_model_spec}")
        lines.append(f"  Equations present: {self.has_equations}")
        lines.append(f"  Standard errors reported: {self.has_standard_errors}")
        lines.append(f"  Variance components reported: {self.has_variance_components}")
        lines.append(f"  Model fit statistics reported: {self.has_model_fit_stats}")
        lines.append(f"  Effect sizes reported: {self.has_effect_sizes}")
        lines.append(f"  Confidence intervals reported: {self.has_confidence_intervals}")
        lines.append(f"  P-values reported: {self.has_p_values}")
        lines.append(f"  Sample/participants described: {self.has_sample_description}")
        lines.append(f"  Randomisation described: {self.has_randomisation}")
        type_counts = {}
        for b in self.blocks:
            type_counts[b.block_type.name] = type_counts.get(b.block_type.name, 0) + 1
        lines.append(f"  Block counts: {type_counts}")
        
        return "\n".join(lines)
        
        revised = revised.replace(
            'labels a',
            'appears to label the standardized effect for a',
        )
        revised = revised.replace(
            'as "trivial" despite the magnitude being potentially',
            'as small or trivial on a standardized scale, although it may still be',
        )


# ---------------------------------------------------------------------------
# Phase 3: Method classification
# ---------------------------------------------------------------------------

class MethodClass(Enum):
    FREQUENTIST_ROBUST = "frequentist_robust"
    FREQUENTIST_REGRESSION = "frequentist_regression"
    FREQUENTIST_ANOVA = "frequentist_anova"
    CROSS_CLASSIFIED_MCMC = "cross_classified_mcmc"
    MULTILEVEL_MCMC = "multilevel_mcmc"
    BAYESIAN_MODEL = "bayesian_model"
    DISTRIBUTIONAL_MODEL = "distributional_model"
    PREDICTIVE_ML = "predictive_ml"
    DESCRIPTIVE = "descriptive"
    MIXED_METHODS = "mixed_methods"
    UNCLASSIFIED = "unclassified"

# Keyword sets for classification — order matters (more specific first)
_METHOD_PATTERNS: List[Tuple[MethodClass, List[str], List[str]]] = [
    # (class, required_any, exclude_any)
    (MethodClass.CROSS_CLASSIFIED_MCMC, [
        r"cross.?classif",
    ], []),

    (MethodClass.DISTRIBUTIONAL_MODEL, [
        r"\bgamlss\b", r"distributional\s+model", r"distributional\s+regression",
        r"\bvgam\b", r"location.?scale.?shape",
    ], []),

    (MethodClass.BAYESIAN_MODEL, [
        r"\bprior[s]?\b.*\bposterior[s]?\b",
        r"\bbrms\b", r"\bstan\b(?!.*\bstan[dt])", r"\bjags\b", r"\bwinbugs\b",
        r"bayes\s*factor", r"\brope\b.*\bposterior\b",
        r"prior\s+predict", r"posterior\s+predict",
        r"\bcredible\s+interval",
    ], []),

    (MethodClass.MULTILEVEL_MCMC, [
        r"\bmlwin\b", r"multilevel.*mcmc", r"hierarchical.*mcmc",
        r"mcmc.*multilevel", r"mcmc.*hierarchical",
        r"random\s+effect.*mcmc", r"mcmc.*random\s+effect",
    ], [
        r"\bprior[s]?\s+(distribut|specif|elicit|sensitiv)",
        r"\bposterior\s+(distribut|densit|summar)",
    ]),

    (MethodClass.FREQUENTIST_ROBUST, [
        r"yuen\s+paired\s*t",
        r"yuen\s+test",
        r"trimmed\s+mean",
        r"trimmed\s+means",
        r"winsori[sz]ed",
        r"\bwrs2\b",
        r"benjamini",
        r"hochberg",
        r"\bs\s+value\b",
        r"surprisal",
        r"robust\s+standardi[sz]ed\s+effect\s+size",
    ], []),

    (MethodClass.PREDICTIVE_ML, [
        r"\bcross.?valid", r"\bROC\b", r"\bAUC\b", r"\bbrier\b",
        r"random\s*forest", r"\bxgboost\b", r"\blasso\b", r"\belastic\s*net\b",
        r"neural\s*net", r"deep\s*learn", r"calibrat.*discriminat",
        r"train.*test\s*split", r"hyperparamet",
    ], []),

    (MethodClass.FREQUENTIST_ANOVA, [
        r"\banova\b", r"\bancova\b", r"\bmanova\b", r"\bmancova\b",
        r"repeated\s*measures", r"factorial\s+design",
        r"between.?subject.*within.?subject",
    ], []),

    (MethodClass.FREQUENTIST_REGRESSION, [
        r"linear\s+regress", r"logistic\s+regress", r"multiple\s+regress",
        r"\bols\b", r"\bglm\b(?!.*gamlss)", r"mixed.?effect.*model",
        r"linear\s+mixed", r"generalised\s+linear",
    ], []),

    (MethodClass.DESCRIPTIVE, [
        r"descriptive\s+statistic", r"descriptive\s+analys",
        r"prevalence\s+stud", r"cross.?sectional\s+surv",
    ], []),
]


def classify_method(full_text: str) -> MethodClass:
    """Rule-based classification of the primary statistical method."""
    text_lower = full_text.lower()

    for method_class, required_patterns, exclude_patterns in _METHOD_PATTERNS:
        # Check exclusions first
        excluded = False
        for pat in exclude_patterns:
            if re.search(pat, text_lower):
                excluded = True
                break
        if excluded:
            continue

        for pat in required_patterns:
            if re.search(pat, text_lower):
                return method_class

    # Fallback: if MCMC/DIC mentioned without Bayesian terms, classify as multilevel MCMC
    if re.search(r"\bmcmc\b", text_lower) and not re.search(r"\bprior[s]?\b", text_lower):
        return MethodClass.MULTILEVEL_MCMC

    return MethodClass.UNCLASSIFIED


# Framework-specific review expectations loaded by method class
_METHOD_EXPECTATIONS: Dict[MethodClass, str] = {
    MethodClass.FREQUENTIST_ROBUST: """Method-specific expectations (robust frequentist):
- Appropriate focus: trimming rationale, robustness to outliers/non-normality, clarity of paired or grouped comparison structure.
- Reporting: trimmed means or robust estimates, interval estimates, effect sizes, multiplicity correction if relevant.
- Prefer concerns about sensitivity to trimming choices, agreement versus association, and whether robust and conventional conclusions diverge.
- Do not default to mixed-model or variance-component expectations unless such models were actually used.
- Do not overemphasise regression diagnostics unless regression is central to the primary inferential claims.
- Do not request Bayesian diagnostics (R-hat, ESS, divergences, priors).
- Do not request prediction metrics unless the paper is explicitly framed as predictive.""",

    MethodClass.FREQUENTIST_REGRESSION: """Method-specific expectations (frequentist regression):
- Appropriate diagnostics: residual plots, multicollinearity checks (VIF where >1 predictor), influence diagnostics, normality of residuals.
- Reporting: coefficients with standard errors or confidence intervals, R-squared or adjusted R-squared, F-test or likelihood ratio test.
- Do not request Bayesian diagnostics (R-hat, ESS, divergences, priors).
- Do not request prediction metrics (AUC, calibration) unless the paper frames it as a prediction study.""",

    MethodClass.FREQUENTIST_ANOVA: """Method-specific expectations (frequentist ANOVA/ANCOVA):
- Appropriate diagnostics: homogeneity of variance (Levene), sphericity (Mauchly for RM), normality, Box's M for MANOVA.
- Reporting: F-values, degrees of freedom, p-values, effect sizes (partial eta-squared, Cohen's d, or similar).
- Post-hoc tests with multiplicity correction if applicable.
- Do not request Bayesian diagnostics or prediction metrics.""",

    MethodClass.CROSS_CLASSIFIED_MCMC: """Method-specific expectations (cross-classified model via MCMC):
- MCMC is an estimation method here, NOT a Bayesian analysis. Treat this as a frequentist multilevel model estimated via MCMC because standard ML/REML cannot handle cross-classified structures efficiently.
- Appropriate diagnostics: convergence (chain mixing, burn-in), DIC or similar information criterion for model comparison, variance component estimates with standard errors.
- Reporting: fixed effect estimates with standard errors, variance components, model comparison statistics (DIC), significance tests (Wald-type chi-square).
- Do NOT request: prior specifications, posterior distributions, credible intervals, Bayes factors, prior sensitivity analyses, R-hat, or any Bayesian reporting elements.
- Do NOT describe the framework as unclear, ambiguous, or hybrid.
- Do NOT request VIF for single-predictor models.
- Do NOT request residual plots as a standard expectation for MCMC-estimated models (convergence diagnostics serve a different role).""",

    MethodClass.MULTILEVEL_MCMC: """Method-specific expectations (multilevel model via MCMC):
- MCMC is an estimation method here, NOT necessarily a Bayesian analysis. If no priors, posteriors, or credible intervals are mentioned, treat as frequentist multilevel modelling using MCMC for estimation.
- Appropriate diagnostics: convergence assessment (burn-in, chain mixing), DIC for model comparison, variance partition coefficients.
- Reporting: fixed effects with standard errors, random effect variance components, model fit statistics.
- Do NOT request Bayesian diagnostics unless explicitly Bayesian language is present in the manuscript.
- Do NOT describe the framework as unclear if MCMC is used without Bayesian terminology.""",

    MethodClass.BAYESIAN_MODEL: """Method-specific expectations (Bayesian model):
- Appropriate diagnostics: prior specification and justification, prior sensitivity analysis, R-hat and effective sample size, divergent transitions (if HMC/NUTS), posterior predictive checks.
- Reporting: posterior summaries (means/medians with credible intervals), prior choices, model comparison (LOO, WAIC, Bayes factors as appropriate).
- Do not request frequentist p-values or confidence intervals.""",

    MethodClass.DISTRIBUTIONAL_MODEL: """Method-specific expectations (distributional/GAMLSS model):
- These models simultaneously model location, scale, and shape parameters. Do not treat scale/shape submodels as unusual.
- Appropriate diagnostics: worm plots or Q-Q plots for distributional fit, randomised quantile residuals, GAIC for model/distribution selection.
- Reporting: coefficients for each distribution parameter, distribution family chosen with justification, model selection criteria.
- Do not request standard OLS residual plots. Do not request VIF unless multiple collinear predictors are modelled.
- Do not treat small coefficients on scale/shape parameters as data errors.""",

    MethodClass.PREDICTIVE_ML: """Method-specific expectations (predictive/ML model):
- Appropriate diagnostics: internal validation (cross-validation, bootstrap), calibration (calibration plots, Hosmer-Lemeshow), discrimination (AUC/C-statistic), decision curve analysis if clinical.
- Reporting: performance metrics with confidence intervals or standard errors, comparison with simpler benchmarks, description of feature engineering and hyperparameter tuning.
- Do not request causal interpretation of feature importance unless the paper frames it causally.""",

    MethodClass.DESCRIPTIVE: """Method-specific expectations (descriptive/epidemiological):
- Appropriate reporting: summary statistics with measures of spread, response rates, sampling strategy, representativeness.
- Do not request inferential diagnostics beyond what the study design supports.""",
}


def get_method_expectations(method_class: MethodClass) -> str:
    """Return framework-specific review expectations, or a generic fallback."""
    return _METHOD_EXPECTATIONS.get(method_class, """Method-specific expectations:
- Method not confidently classified. Apply general standards: appropriate diagnostics for the analysis type, clear reporting of estimates and uncertainty, and model assumptions addressed.
- Match your diagnostic expectations to the actual analysis used. Do not default to OLS assumptions.""")


# ---------------------------------------------------------------------------
# Evidence block classification (heuristic, not LLM)
# ---------------------------------------------------------------------------

# Patterns that signal specific block types
_TABLE_SIGNALS = re.compile(
    r"\[TABLE_START\]|^\s*\|.+\|.+\|", re.MULTILINE
)
_MODEL_SIGNALS = re.compile(
    r"(?i)(?:"
    r"model\s+equation|regression\s+model|fitted\s+model|model\s+specification"
    r"|(?:fixed|random)\s+effect|variance\s+component"
    r"|coefficient.*estimate|~\s*N\(|\\?beta_?\d"
    r"|equation\s+\d|eq\.\s*\d|formula:"
    r"|AIC\b|BIC\b|DIC\b|WAIC\b|LOO\b"
    r"|R[\-\s]?squared|adjusted\s+R|deviance"
    r"|logit\(|log\(|link\s+function"
    r"|cross[\-\s]?classif.*model|multilevel\s+model|hierarchical\s+model"
    r"|MCMC.*model|model.*MCMC"
    r"|between[\-\s]?\w+\s+variance|within[\-\s]?\w+\s+variance"
    r")"
)
_DESIGN_SIGNALS = re.compile(
    r"(?i)(?:"
    # Narrowed: removed generic terms like "condition", "procedure", "protocol"
    # that appear in results and discussion sections too
    r"participant[s]?\s+(?:were|gave|volunteer|complet|recruit)"
    r"|sample\s+size|recruit(?:ed|ment)|volunteer(?:ed|s)"
    r"|informed\s+consent|written\s+consent"
    r"|inclusion\s+criter|exclusion\s+criter|eligib"
    r"|randomi[sz](?:ed|ation)\s+(?:to|into|between)"
    r"|control\s+group|experimental\s+group"
    r"|between[\-\s]?subject.*design|within[\-\s]?subject.*design"
    r"|counterbalance"
    r"|ethical\s+approv|IRB|ethics\s+committee"
    r"|materials?\s+and\s+methods\b"
    r")"
)
_RESULT_SIGNALS = re.compile(
    r"(?i)(?:"
    r"p\s*[<=<]\s*0?\.\d|p\s*=\s*0?\.\d"
    r"|chi[\-\s]?square|chi2|χ2|χ²"
    r"|F\s*\(\s*\d|t\s*\(\s*\d|z\s*=\s*[\-\d]"
    r"|cohen['\u2019]?s\s*d|eta[\-\s]?squared|partial\s+eta"
    r"|odds\s+ratio|hazard\s+ratio|risk\s+ratio|relative\s+risk"
    r"|confidence\s+interval|CI\s*[=:\[]|95\s*%\s*CI"
    r"|standard\s+error|SE\s*=|SD\s*="
    r"|mean\s+difference|effect\s+size"
    r"|statistically\s+significant|non[\-\s]?significant"
    r"|β\s*=|b\s*=\s*[\-\d]"
    r"|the\s+results?\s+(?:show|indicat|suggest|support|reveal)"
    r"|difference\s+of\s+\d"
    r"|practical(?:ly)?\s+significant"
    r")"
)
_FIGURE_CAPTION_SIGNALS = re.compile(
    r"(?i)(?:^|\n)\s*(?:Fig(?:ure)?\.?\s+\d|FIGURE\s+\d)", re.MULTILINE
)

# Section headings that override paragraph-level classification
_SECTION_HEADING_HINTS = {
    BlockType.RESULT: re.compile(
        r"(?i)(?:^|\n)\s*(?:RESULTS?\b|FINDINGS?\b|CROSS[\-\s]?CLASSIFIED\s+MODEL"
        r"|PRACTICAL\s+SIGNIFICANCE|STATISTICAL\s+ANALYSIS)",
    ),
    BlockType.MODEL: re.compile(
        r"(?i)(?:^|\n)\s*(?:ANALYSIS\b|STATISTICAL\s+MODEL|MODEL\s+SPECIFICATION"
        r"|MODELL?ING\s+APPROACH|CROSS[\-\s]?CLASSIFIED\s+(?:MODEL|ANALYSIS))",
    ),
    BlockType.DESIGN: re.compile(
        r"(?i)(?:^|\n)\s*(?:MATERIALS?\s+AND\s+METHODS\b|METHOD(?:S|OLOGY)\b"
        r"|PARTICIPANTS?\b|PROCEDURE\b|STUDY\s+DESIGN\b|DATA\s+COLLECTION\b)",
    ),
}


def classify_paragraph_block(text: str, page: Optional[int] = None) -> BlockType:
    """Classify a paragraph-level text segment into a block type."""
    # TABLE_START marker is definitive
    if _TABLE_SIGNALS.search(text):
        return BlockType.TABLE

    # Check for section heading overrides in the first 120 characters
    head = text[:120]
    for block_type, pattern in _SECTION_HEADING_HINTS.items():
        if pattern.search(head):
            return block_type

    if _FIGURE_CAPTION_SIGNALS.search(text) and len(text) < 600:
        return BlockType.FIGURE_CAPTION

    # Count signal density for competing types, with weighting
    model_hits = len(_MODEL_SIGNALS.findall(text))
    design_hits = len(_DESIGN_SIGNALS.findall(text))
    result_hits = len(_RESULT_SIGNALS.findall(text))

    # MODEL and RESULT signals are weighted more heavily because DESIGN
    # terms (participant, condition, procedure) appear across all sections,
    # while model specifications and test statistics are more localised
    scores = {
        BlockType.MODEL: model_hits * 2.0,
        BlockType.DESIGN: design_hits * 1.0,
        BlockType.RESULT: result_hits * 1.5,
    }
    best = max(scores, key=scores.get)
    if scores[best] >= 2:
        return best

    return BlockType.NARRATIVE


def structure_evidence(
    source_name: str,
    text: str,
    table_blocks: List[Tuple[int, str]],
) -> EvidenceManifest:
    """
    Convert raw extracted text and table blocks into typed EvidenceBlocks
    and build a manifest summarising what evidence is available.
    """
    manifest = EvidenceManifest(source_name=source_name)
    text_lower = text.lower()

    # --- Structured table blocks (from pdfplumber/camelot) ---
    for page_num, block_text in table_blocks:
        label_match = re.search(r"Label:\s*(.+)", block_text)
        label = label_match.group(1).strip() if label_match else None
        manifest.blocks.append(EvidenceBlock(
            block_type=BlockType.TABLE,
            text=block_text,
            source_name=source_name,
            page=page_num,
            label=label,
        ))
        if label:
            manifest.table_labels.append(label)

    # --- Inline table recovery from prose text ---
    inline_tables = detect_inline_tables(text)
    for itbl in inline_tables:
        # Avoid duplicating tables already captured structurally
        if not _inline_table_duplicates_existing(itbl["text"], manifest.blocks):
            manifest.blocks.append(EvidenceBlock(
                block_type=BlockType.TABLE,
                text=itbl["text"],
                source_name=source_name,
                label=itbl.get("label"),
                confidence="medium",
            ))
            if itbl.get("label"):
                manifest.table_labels.append(itbl["label"])

    manifest.n_tables = len([b for b in manifest.blocks if b.block_type == BlockType.TABLE])

    # --- Paragraph-level classification ---
    paragraphs = text.split("\n\n")
    for para in paragraphs:
        para = para.strip()
        if not para or len(para) < 30:
            continue
        block_type = classify_paragraph_block(para)
        # Do not duplicate TABLE blocks already captured
        if block_type == BlockType.TABLE and "[TABLE_START]" in para:
            continue
        manifest.blocks.append(EvidenceBlock(
            block_type=block_type,
            text=para,
            source_name=source_name,
        ))

    # --- Manifest flags from full text ---
    manifest.has_equations = bool(re.search(
        r"(?:equation\s+\d|eq\.\s*\d|~\s*N\(|Score.*=.*β|model\s+equation)", text_lower
    ))
    manifest.has_model_spec = bool(re.search(
        r"(?:model\s+equation|regression\s+model|cross.?classif|fitted\s+model|model\s+specification)", text_lower
    ))
    manifest.has_effect_sizes = bool(re.search(
        r"(?:cohen|eta.?squared|partial\s+eta|effect\s+size|odds\s+ratio|hazard\s+ratio)", text_lower
    ))
    manifest.has_confidence_intervals = bool(re.search(
        r"(?:confidence\s+interval|95\s*%\s*ci|\bci\s*[=:\[])", text_lower
    ))
    manifest.has_p_values = bool(re.search(
        r"(?:p\s*[<=<]\s*0?\.\d|p\s*=\s*0?\.\d)", text_lower
    ))
    manifest.has_sample_description = bool(re.search(
        r"(?:participant|sample\s+size|n\s*=\s*\d|volunteer|recruit)", text_lower
    ))
    manifest.has_randomisation = bool(re.search(
        r"(?:randomi[sz]|random\s+allocat|random\s+assign)", text_lower
    ))
    manifest.has_standard_errors = bool(re.search(
        r"(?:standard\s+error|SE\s*=|s\.e\.\s*=|\(\s*\d+\.\d+\s*\)\s*$)", text_lower
    ))
    manifest.has_variance_components = bool(re.search(
        r"(?:variance\s+component|between.?\w+\s+variance|random.*variance|σ²|sigma\s*squared)", text_lower
    ))
    manifest.has_model_fit_stats = bool(re.search(
        r"(?:\bDIC\b|\bAIC\b|\bBIC\b|\bWAIC\b|\bLOO\b|deviance\s+information|"
        r"R[\-\s]?squared|adjusted\s+R|log[\-\s]?likelihood|-2\s*log)", text_lower
    ))
    manifest.n_figures_mentioned = len(re.findall(
        r"(?i)(?:fig(?:ure)?\.?\s+\d)", text
    ))

    # --- Method classification ---
    manifest.method_class = classify_method(text)
    
    combined_table_text = "\n\n".join(block_text for _, block_text in table_blocks)
    combined_block_text = "\n\n".join(b.text for b in manifest.blocks if b.block_type == BlockType.TABLE)
    full_evidence_text = text + "\n\n" + combined_table_text + "\n\n" + combined_block_text
    manifest = refine_manifest_flags(manifest, full_evidence_text)

    return manifest


def detect_inline_tables(text: str) -> List[dict]:
    """
    Detect tabular data embedded in prose text, including:
    1. Markdown pipe tables from Marker output
    2. Numeric column patterns (rows of aligned numbers)
    3. Table headers followed by numeric rows
    """
    tables: List[dict] = []

    # --- Pattern 1: Markdown pipe tables (from Marker) ---
    pipe_table_re = re.compile(
        r"((?:^[ \t]*\|.+\|[ \t]*\n){3,})",
        re.MULTILINE,
    )
    for m in pipe_table_re.finditer(text):
        block_text = m.group(1).strip()
        # Skip separator-only tables
        rows = [r for r in block_text.split("\n") if r.strip() and not re.match(r"^\s*\|[\s\-:]+\|\s*$", r)]
        if len(rows) >= 2:
            label = _infer_inline_table_label(text, m.start())
            tables.append({
                "text": f"[TABLE_START]\nSource: inline_markdown\n{block_text}\n[TABLE_END]",
                "label": label,
            })

    # --- Pattern 2: Rows of aligned numbers preceded by a header ---
    # Matches blocks like:
    #   Score to home | No noise | Noise | Total
    #   −4.00           6          4       10
    #   −3.00           5          4       9
    # Look for 3+ consecutive lines with ≥2 numbers each
    lines = text.split("\n")
    i = 0
    while i < len(lines):
        run_start = None
        run_lines = []
        j = i
        while j < len(lines):
            line = lines[j].strip()
            # Count numeric tokens (integers, decimals, negatives)
            nums = re.findall(r"-?\d+\.?\d*", line)
            if len(nums) >= 2 and len(line.split()) <= 12:
                if run_start is None:
                    run_start = j
                run_lines.append(line)
            else:
                if run_lines and len(run_lines) >= 3:
                    break
                run_start = None
                run_lines = []
            j += 1

        if run_lines and len(run_lines) >= 3:
            # Include a header line if the line before the run has words
            header = ""
            if run_start and run_start > 0:
                candidate = lines[run_start - 1].strip()
                if candidate and re.search(r"[A-Za-z]", candidate):
                    header = candidate + "\n"

            block_text = header + "\n".join(run_lines)
            label = _infer_inline_table_label(text, text.find(run_lines[0]))
            tables.append({
                "text": f"[TABLE_START]\nSource: inline_numeric\n{block_text}\n[TABLE_END]",
                "label": label,
            })
            i = j
        else:
            i = j + 1 if j > i else i + 1

    return tables


def _infer_inline_table_label(text: str, position: int) -> Optional[str]:
    """Look backwards from a table position for a Table N label."""
    preceding = text[max(0, position - 300):position]
    m = re.search(r"(Table\s+\d+[a-z]?)\b", preceding, re.IGNORECASE)
    return m.group(1) if m else None


def _inline_table_duplicates_existing(inline_text: str, existing_blocks: List[EvidenceBlock]) -> bool:
    """Check whether an inline table duplicates data already in a structured block."""
    # Extract the numeric values from the inline table
    inline_nums = set(re.findall(r"-?\d+\.?\d*", inline_text))
    if len(inline_nums) < 4:
        return False
    for block in existing_blocks:
        if block.block_type == BlockType.TABLE:
            block_nums = set(re.findall(r"-?\d+\.?\d*", block.text))
            # If >70% of inline numbers appear in an existing block, it is a duplicate
            if inline_nums and len(inline_nums & block_nums) / len(inline_nums) > 0.7:
                return True
    return False


# ---------------------------------------------------------------------------
# Phase 6: Programmatic post-validation
# ---------------------------------------------------------------------------

def programmatic_post_checks(report_text: str, manifest: EvidenceManifest) -> List[str]:
    """
    Catch common LLM reviewer errors by comparing claims in the report
    against what the evidence manifest confirms is present.

    Returns a list of correction instructions to feed into LLM validation.
    """
    corrections: List[str] = []
    report_lower = report_text.lower()

    # 1. "Table missing" when tables are present
    if manifest.n_tables > 0:
        if re.search(r"(?:table[s]?\s+(?:are|is|were|was)\s+(?:missing|absent|not\s+provided|not\s+reported))", report_lower):
            corrections.append(
                f"CORRECTION: The report claims tables are missing, but {manifest.n_tables} table(s) were extracted "
                f"({', '.join(manifest.table_labels) if manifest.table_labels else 'unlabelled'}). "
                f"Remove or revise this claim."
            )

    # 2. "Equation/model spec missing" when present
    if manifest.has_model_spec or manifest.has_equations:
        if re.search(r"(?:equation|model\s+(?:specification|form|equation))\s+(?:is|are|was|were)\s+(?:missing|absent|not\s+provided|not\s+reported|unclear)", report_lower):
            corrections.append(
                "CORRECTION: The report claims the model specification or equation is missing/unclear, "
                "but the extracted evidence contains model specification content. Revise to acknowledge this."
            )

    # 3. Requesting VIF for single-predictor models
    if re.search(r"\bvif\b", report_lower):
        model_blocks = [b for b in manifest.blocks if b.block_type == BlockType.MODEL]
        all_model_text = " ".join(b.text for b in model_blocks).lower()
        # Also check the full text for single-predictor indicators
        full_check = all_model_text + " " + " ".join(
            b.text.lower() for b in manifest.blocks
            if b.block_type in (BlockType.RESULT, BlockType.DESIGN)
        )
        if re.search(r"single\s+fixed\s+(?:factor|effect|predictor)|single\s+predictor|one\s+predictor|single\s+.*\bfactor\b", full_check):
            corrections.append(
                "CORRECTION: The report requests VIF checks, but the model has a single predictor. "
                "VIF is not applicable. Remove this request."
            )

    # 4. Requesting Bayesian diagnostics for non-Bayesian MCMC models
    mc = manifest.method_class
    if mc in (MethodClass.CROSS_CLASSIFIED_MCMC, MethodClass.MULTILEVEL_MCMC):
        bayesian_requests = re.findall(
            r"(?:prior\s+(?:specification|sensitivity|justification|choice)"
            r"|posterior\s+(?:distribution|summary|check)"
            r"|credible\s+interval"
            r"|bayes\s+factor"
            r"|framework\s+(?:is|remains|appears)\s+(?:unclear|ambiguous|hybrid))",
            report_lower
        )
        if bayesian_requests:
            corrections.append(
                f"CORRECTION: The model uses MCMC as an estimation method (classified as {mc.value}), "
                f"not as a Bayesian analysis. The report contains {len(bayesian_requests)} reference(s) to "
                f"Bayesian diagnostics or calls the framework unclear. Remove all such references."
            )

    # 5. Requesting OLS-style residual plots for MCMC/Bayesian/GAMLSS models
    if mc in (MethodClass.CROSS_CLASSIFIED_MCMC, MethodClass.MULTILEVEL_MCMC,
              MethodClass.BAYESIAN_MODEL, MethodClass.DISTRIBUTIONAL_MODEL):
        if re.search(r"residual\s+plot", report_lower):
            corrections.append(
                f"CORRECTION: The report requests residual plots, but the analysis uses {mc.value}. "
                f"Standard OLS residual plots are not the primary diagnostic for this framework. "
                f"Revise to request framework-appropriate diagnostics (e.g., convergence diagnostics for MCMC, "
                f"randomised quantile residuals for GAMLSS) or remove if not applicable."
            )

    # 6. "Sample description missing" when present
    if manifest.has_sample_description:
        if re.search(r"sample\s+(?:size|description)\s+(?:is|are|was|were)\s+(?:missing|not\s+reported|absent)", report_lower):
            corrections.append(
                "CORRECTION: The report claims sample description is missing, but participant/sample "
                "information was found in the extracted evidence. Revise this claim."
            )

    # 7. "Randomisation unclear" when described
    if manifest.has_randomisation:
        if re.search(r"randomi[sz]ation\s+(?:is|was|remains)\s+(?:unclear|ambiguous|not\s+described|not\s+specified)", report_lower):
            corrections.append(
                "CORRECTION: The report claims randomisation is unclear or not described, but "
                "randomisation details were found in the extracted evidence. Revise: the criticism "
                "should target specific aspects that remain unclear, not claim absence."
            )

    # 8. "P-values/CIs not reported" when present
    if manifest.has_p_values:
        if re.search(r"p[\-\s]?value[s]?\s+(?:are|is|were|was)\s+(?:not\s+reported|missing|absent)", report_lower):
            corrections.append(
                "CORRECTION: The report claims p-values are not reported, but p-values were found "
                "in the extracted evidence. Remove this claim."
            )

    if manifest.has_confidence_intervals:
        if re.search(r"confidence\s+interval[s]?\s+(?:are|is|were|was)\s+(?:not\s+reported|missing|absent)", report_lower):
            corrections.append(
                "CORRECTION: The report claims confidence intervals are not reported, but CIs were "
                "found in the extracted evidence. Remove this claim."
            )

    # 9. "Standard errors not reported" when present
    if manifest.has_standard_errors:
        if re.search(r"standard\s+error[s]?\s+(?:are|is|were|was)\s+(?:not\s+reported|missing|absent)", report_lower):
            corrections.append(
                "CORRECTION: The report claims standard errors are not reported, but standard errors "
                "were found in the extracted evidence. Remove this claim."
            )

    # 10. "Variance components not reported/tested" when present
    if manifest.has_variance_components:
        if re.search(r"variance\s+component[s]?\s+(?:are|is|were|was)\s+(?:not\s+reported|missing|absent)", report_lower):
            corrections.append(
                "CORRECTION: The report claims variance components are not reported, but variance "
                "component estimates were found in the extracted evidence. Revise this claim."
            )

    # 11. "Model fit not reported" when DIC/AIC/BIC is present
    if manifest.has_model_fit_stats:
        if re.search(r"(?:model\s+fit|goodness[\-\s]of[\-\s]fit|fit\s+statistic)[s]?\s+(?:are|is|were|was)\s+(?:not\s+reported|missing|absent)", report_lower):
            corrections.append(
                "CORRECTION: The report claims model fit statistics are not reported, but fit "
                "statistics (e.g., DIC, AIC, deviance) were found in the extracted evidence."
            )

    # 12. Cross-contamination: numbers from cited studies confused with current study
    # Detect when the report claims a sample discrepancy using a number that
    # belongs to a different study cited in the text
    contamination_issues = _detect_cross_contamination(report_lower, manifest)
    corrections.extend(contamination_issues)

    return corrections
    
def refine_manifest_flags(manifest: dict, combined_text: str, table_blocks: list[str]) -> dict:
    text_low = combined_text.lower()
    tables_low = "\n\n".join(table_blocks).lower()

    def seen(*patterns: str) -> bool:
        return any(p.lower() in text_low or p.lower() in tables_low for p in patterns)

    # model specification / equations
    if seen("the following equation", "equation was applied", "ln(vo2max", "p = w", "work–time model", "work-time model"):
        manifest["model_spec_present"] = True

    if seen("eq.", "equation", "ln(", "exp(", "vo2max(", "p = w"):
        manifest["equations_present"] = True

    # standard errors / SDs
    if seen("\tse\t", " se ", "standard error", "trimmed sd", "sd (0.2)", "winsorized variances"):
        manifest["standard_errors_reported"] = True

    # fit stats / performance stats
    if seen("r squared", "adjusted r squared", "aic", "bic", "looic", "dic", "rmse", "mae", "effect size", "s value", "surprisal"):
        manifest["model_fit_statistics_reported"] = True
        manifest["effect_sizes_reported"] = True

    # confidence intervals
    if seen("95% ci", "confidence interval", "lower bound", "upper bound", "\tlower\tupper\t", "lower upper"):
        manifest["confidence_intervals_reported"] = True

    # p-values
    if seen("p =", "p <", "p ≤", "sig.", "benjamini", "hochberg"):
        manifest["p_values_reported"] = True

    # randomisation
    if seen("randomized", "randomised", "participants were randomized", "order of test completion"):
        manifest["randomisation_described"] = True

    return manifest
    
def refine_manifest_flags(manifest: EvidenceManifest, full_text: str) -> EvidenceManifest:

    """
    Upgrade manifest flags when evidence is clearly present in text or table blocks
    but was missed by the initial coarse regex checks.
    """
    full_low = full_text.lower()

    def seen(*patterns: str) -> bool:
        return any(p.lower() in full_low for p in patterns)

    # Model specification / equations
    if seen("the following equation", "equation was applied", "model equation", "ln(vo2max", "p = w", "work–time model", "work-time model"):
        manifest.has_model_spec = True
        manifest.has_equations = True

    # Confidence intervals
    if seen("95% ci", "confidence interval", "lower bound", "upper bound", " lower ", " upper "):
        manifest.has_confidence_intervals = True

    # Standard errors / SD-like reporting
    if seen(" se ", "\tse\t", "standard error", "trimmed sd", "sd (0.2)", "winsorized variances"):
        manifest.has_standard_errors = True

    # Fit statistics / performance metrics / effect sizes
    if seen("r squared", "adjusted r squared", "aic", "bic", "waic", "loo", "looic", "dic", "rmse", "mae", "effect size", "s value", "surprisal"):
        manifest.has_model_fit_stats = True

    if seen("effect size", "cohen", "robust version of cohen", "eta squared", "partial eta", "odds ratio", "hazard ratio", "s value"):
        manifest.has_effect_sizes = True

    # P-values
    if seen("p =", "p <", "p ≤", "sig.", "benjamini", "hochberg"):
        manifest.has_p_values = True

    # Randomisation
    if seen("randomized", "randomised", "randomized into", "randomised into", "randomization", "randomisation", "order of test completion"):
        manifest.has_randomisation = True

    return manifest

def _detect_cross_contamination(report_lower: str, manifest: EvidenceManifest) -> List[str]:
    """
    Detect when the LLM has confused numbers from cited studies with the
    current study's data, creating spurious discrepancy claims.
    """
    corrections: List[str] = []

    # Look for "discrepancy" or "unexplained" claims involving specific numbers
    discrepancy_patterns = re.findall(
        r"(?:unexplained|discrepan|inconsisten|contradict)\w*"
        r".{0,80}?"  # allow up to 80 chars between keyword and first number
        r"(\d{2,})"
        r".{0,40}?"  # allow up to 40 chars between first number and comparison word
        r"(?:vs\.?|versus|and|compared\s+to|but|yet|however)"
        r".{0,40}?"  # allow up to 40 chars between comparison word and second number
        r"(\d{2,})",
        report_lower,
    )

    if not discrepancy_patterns:
        return corrections

    # Gather numbers that appear in citation contexts (near "et al.", year references,
    # or phrases indicating a different study)
    all_text = " ".join(b.text for b in manifest.blocks)
    cited_numbers: set = set()

    # Match numbers near citation markers (et al., year references)
    for m in re.finditer(r"(?:et\s+al\.?\s*[\(,]?\s*\d{4}\)?|[\(]\d{4}[\)])", all_text):
        context_start = max(0, m.start() - 200)
        context_end = min(len(all_text), m.end() + 200)
        context = all_text[context_start:context_end]
        for num in re.findall(r"\b(\d{2,})\b", context):
            cited_numbers.add(num)

    # Also flag numbers near "previous study", "larger dataset", "earlier study" etc.
    for m in re.finditer(r"(?:previous\s+\w*\s*(?:study|dataset|research|investigation)"
                         r"|larger\s+dataset|earlier\s+\w*\s*study|separate\s+\w*\s*study"
                         r"|prior\s+\w*\s*(?:study|research))", all_text, re.IGNORECASE):
        context_start = max(0, m.start() - 200)
        context_end = min(len(all_text), m.end() + 200)
        context = all_text[context_start:context_end]
        for num in re.findall(r"\b(\d{2,})\b", context):
            cited_numbers.add(num)

    for num1, num2 in discrepancy_patterns:
        if num1 in cited_numbers or num2 in cited_numbers:
            corrections.append(
                f"CORRECTION: The report identifies a discrepancy between {num1} and {num2}, but "
                f"at least one of these numbers appears in the context of a cited study (e.g., "
                f"a different dataset referenced by the authors), not the current study's data. "
                f"Verify that this is a genuine within-study discrepancy before claiming inconsistency. "
                f"If the numbers come from different studies, remove the discrepancy claim."
            )
            break  # One correction is sufficient to flag the issue

    return corrections


# ---------------------------------------------------------------------------
# OCR and text cleaning (preserved from original)
# ---------------------------------------------------------------------------

SUPERSCRIPT_MAP = str.maketrans({
    "0": "⁰", "1": "¹", "2": "²", "3": "³", "4": "⁴",
    "5": "⁵", "6": "⁶", "7": "⁷", "8": "⁸", "9": "⁹",
    "+": "⁺", "-": "⁻", "=": "⁼", "(": "⁽", ")": "⁾",
    "n": "ⁿ", "i": "ⁱ",
})

SUBSCRIPT_MAP = str.maketrans({
    "0": "₀", "1": "₁", "2": "₂", "3": "₃", "4": "₄",
    "5": "₅", "6": "₆", "7": "₇", "8": "₈", "9": "₉",
    "+": "₊", "-": "₋", "=": "₌", "(": "₍", ")": "₎",
    "a": "ₐ", "e": "ₑ", "i": "ᵢ", "j": "ⱼ", "o": "ₒ",
    "r": "ᵣ", "u": "ᵤ", "v": "ᵥ", "x": "ₓ",
})


def _to_superscript(s: str) -> str:
    return s.translate(SUPERSCRIPT_MAP)


def _to_subscript(s: str) -> str:
    return s.translate(SUBSCRIPT_MAP)


def clean_ocr_artifacts(text: str) -> str:
    """Centralised function to fix PDF OCR corruption across both text and tables."""
    text = text.replace("\x00", " ").replace("", "")
    text = text.replace("(cid:3)", "-").replace("(cid:2)", "±")

    # Fix the specific index name
    text = text.replace("WHT$5R", "WHT.5R")
    text = text.replace("WHT$5", "WHT.5")

    # Fix the $ replacing decimal points in numbers (e.g., 0$249 -> 0.249, 177$5 -> 177.5)
    text = re.sub(r"(\d)\$(\d)", r"\1.\2", text)

    # Fix OCR 'Z' being used instead of '='
    text = re.sub(r"\bWHT\.5R Z\b", "WHT.5R =", text)
    text = re.sub(r"\br Z\b", "r =", text)
    text = re.sub(r"\bF Z\b", "F =", text)
    text = re.sub(r"\bp Z\b", "p =", text)

    return text


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("", "")
    # Note: Using " +" instead of "[ \t]+" to preserve tabs for table columns
    text = re.sub(r" +", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clean_extracted_evidence_text(text: str, domain: str = "general") -> str:
    text = clean_text(text)
    text = clean_ocr_artifacts(text)

    replacements = {
        "VO 2max": "VO2max",
        "VO 2peak": "VO2peak",
        "VO ₂max": "VO2max",
        "VO ₂peak": "VO2peak",
        "ln (": "ln(",
        "Ln (": "ln(",
        "Age 2": "Age2",
        "age 2": "age2",
        "L. min−1": "L·min−1",
        "mL. kg−1.min−1": "mL·kg−1·min−1",
        "kg −1": "kg−1",
        "min −1": "min−1",
        "body fat %": "body fat%",
        "BF %": "BF%",
        "Cyc le": "Cycle",
        "algor ithms": "algorithms",
        "method- ology": "methodology",
        "allome- tric": "allometric",
        "physiolog- ical": "physiological",
        "incor - porating": "incorporating",
        "log- transformed": "log-transformed",
        "out- of- sample": "out-of-sample",
        "all- cause": "all-cause",
        "fat- free": "fat-free",
        "whole- body": "whole-body",
        "body- composition": "body-composition",
        "X- ray": "X-ray",
        "meta- analysis": "meta-analysis",
        "cross- validation": "cross-validation",
        "peer- reviewed": "peer-reviewed",
        "non- linear": "non-linear",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"([A-Za-z])-\s+([A-Za-z])", r"\1\2", text)
    text = re.sub(r"ln\(mass\(\s*g\s*\)\)", "ln(mass)", text, flags=re.IGNORECASE)
    text = re.sub(r"ln\(mass\s*\(\s*g\s*\)\)", "ln(mass)", text, flags=re.IGNORECASE)
    text = re.sub(r"\bVO\s*\(?\s*2max\s*\)?", "VO2max", text)

    if domain == "exercise_physiology":
        text = re.sub(r"\bln\(masskg\)\b", "ln(mass kg)", text, flags=re.IGNORECASE)
        text = re.sub(r"\bln\(fffkg\)\b", "ln(FFM kg)", text, flags=re.IGNORECASE)
        text = re.sub(r"\blnfvc\b", "ln(FVC)", text, flags=re.IGNORECASE)
        text = re.sub(r"\blnwc\b", "ln(WC)", text, flags=re.IGNORECASE)

    text = re.sub(r" +", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ---------------------------------------------------------------------------
# Table extraction (preserved from original)
# ---------------------------------------------------------------------------

@dataclass
class DocChunk:
    source_name: str
    chunk_id: int
    text: str


def truncate_cell(value, max_chars: int = MAX_TABLE_CHARS_PER_CELL) -> str:
    if value is None:
        return ""
    s = str(value).strip().replace("\n", " ")
    s = re.sub(r" {2,}", " ", s)
    if len(s) > max_chars:
        s = s[: max_chars - 3] + "..."
    return s


def table_to_block(table: List[List[str]], table_num: int, page_num: int) -> str:
    if not table:
        return ""

    trimmed_rows = table[:MAX_TABLE_ROWS]
    lines = [
        "[TABLE_START]",
        f"Page: {page_num}",
        f"Table number on page: {table_num}",
    ]
    for row in trimmed_rows:
        row = row[:MAX_TABLE_COLS]
        cleaned = [truncate_cell(cell) for cell in row]
        lines.append("\t".join(cleaned))
    lines.append("[TABLE_END]")
    return "\n".join(lines).strip()


def normalise_table_rows(rows: list[list[str]]) -> list[list[str]]:
    cleaned = []
    for row in rows:
        cleaned_row = []
        for cell in row:
            cell = "" if cell is None else str(cell)
            cell = clean_ocr_artifacts(cell)
            cell = re.sub(r" +", " ", cell).strip()
            cleaned_row.append(cell)
        if any(cell for cell in cleaned_row):
            cleaned.append(cleaned_row)
    return cleaned


def extract_tables_pdfplumber(path: Path) -> list[dict]:
    tables = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                try:
                    raw_tables = page.extract_tables()
                except Exception:
                    raw_tables = []
                table_counter = 0
                for tbl in raw_tables or []:
                    if not tbl:
                        continue
                    rows = normalise_table_rows(tbl)
                    if len(rows) < 2:
                        continue
                    table_counter += 1
                    tables.append({
                        "source": "pdfplumber",
                        "page": page_index,
                        "table_index": table_counter,
                        "label": infer_table_label({"rows": rows}),
                        "rows": rows,
                    })
    except Exception:
        pass
    return tables


def extract_tables_camelot(path: Path) -> list[dict]:
    tables = []
    try:
        import camelot
    except Exception:
        return tables

    for flavor in ("lattice", "stream"):
        try:
            found = camelot.read_pdf(str(path), pages="all", flavor=flavor)
        except Exception:
            continue

        for i, tbl in enumerate(found, start=1):
            try:
                df = tbl.df
            except Exception:
                continue

            if df is None or df.empty:
                continue

            rows = normalise_table_rows(df.fillna("").values.tolist())
            if len(rows) < 2:
                continue

            page_value = getattr(tbl, "page", None)
            tables.append({
                "source": f"camelot_{flavor}",
                "page": page_value,
                "table_index": i,
                "label": infer_table_label({"rows": rows}),
                "rows": rows,
            })
    return tables


def infer_table_label(table: dict) -> str | None:
    rows = table.get("rows", [])
    head_text = " ".join(" ".join(str(c) for c in row if str(c).strip()) for row in rows[:6])
    m = re.search(r"\b(Table\s+[A-Za-z0-9]+[a-z]?)\b", head_text, flags=re.I)
    if m:
        label = m.group(1)
        label = re.sub(r" +", " ", label).strip()
        return label
    return None


def row_is_prose_like(row: list[str]) -> bool:
    text = " ".join(str(c) for c in row if str(c).strip()).strip()
    if not text:
        return False
    word_count = len(text.split())
    digit_count = len(re.findall(r"\d", text))
    return word_count >= 10 and digit_count < 2


def _is_reference_list(rows: list[list[str]]) -> bool:
    """Detect when a 'table' extraction is actually a reference list."""
    all_text = " ".join(
        " ".join(str(c) for c in row) for row in rows
    ).lower()
    # Reference lists contain many author-year patterns and journal names
    ref_markers = len(re.findall(
        r"(?:\(\d{4}\)|\b\d{4}\b[,.\s]|et\s+al|j\.\s+\w+\s+\w+|psychol\.|sci\.|exerc\.)",
        all_text,
    ))
    # If more than 8 reference-like markers, this is a reference list
    if ref_markers > 8:
        return True
    # Also check for "REFERENCES" heading
    if re.search(r"\breferences\b", all_text[:200]):
        return True
    return False


def _count_numeric_cells(rows: list[list[str]]) -> int:
    """Count cells that contain meaningful numeric content."""
    count = 0
    for row in rows:
        for cell in row:
            s = str(cell).strip()
            # A numeric cell contains at least one digit and is not pure prose
            if s and re.search(r"\d", s) and len(s.split()) <= 5:
                count += 1
    return count


def score_table_candidate(table: dict) -> int:
    rows = table.get("rows", [])
    nonempty_cells = sum(1 for row in rows for cell in row if str(cell).strip())
    nrows = len(rows)
    ncols = max((len(r) for r in rows), default=0)

    prose_like_rows = sum(1 for row in rows if row_is_prose_like(row))
    prose_ratio = prose_like_rows / max(nrows, 1)
    numeric_cells = _count_numeric_cells(rows)

    # --- Reject garbage extractions ---
    # Too few meaningful cells (e.g., pdfplumber returning just "12" and "5")
    if nonempty_cells < 6:
        return -9999
    # Too few numeric cells for a data table
    if numeric_cells < 3 and nrows > 1:
        return -9999
    # Reference list false positive
    if _is_reference_list(rows):
        return -9999

    if table.get("source") == "camelot_stream":
        if prose_ratio > 0.15:
            return -9999
        if ncols < 3 and nonempty_cells < 15:
            return -9999

    # General prose-ratio filter for all sources
    if prose_ratio > 0.5:
        return -9999

    score = nonempty_cells + 5 * nrows + 3 * ncols + 2 * numeric_cells

    source = table.get("source", "")
    if source == "pdfplumber":
        score += 15
    elif source == "camelot_lattice":
        score += 10
    elif source == "camelot_stream":
        score += 5

    label = infer_table_label(table)
    if label:
        score += 20
        table["label"] = label

    return score


def deduplicate_and_select_best_tables(table_candidates: list[dict]) -> list[dict]:
    filtered = []
    for table in table_candidates:
        score = score_table_candidate(table)
        if score > 0:
            if not table.get("label"):
                table["label"] = infer_table_label(table)
            filtered.append(table)

    grouped: dict[tuple, list[dict]] = {}
    unlabelled_by_page: dict[int, list[dict]] = {}

    for table in filtered:
        page = table.get("page")
        label = table.get("label")
        if label:
            key = (page, label)
            grouped.setdefault(key, []).append(table)
        else:
            if page is not None:
                unlabelled_by_page.setdefault(page, []).append(table)

    selected = []
    for _, candidates in grouped.items():
        best = max(candidates, key=score_table_candidate)
        selected.append(best)

    selected_pages = {t.get("page") for t in selected}
    for page, candidates in unlabelled_by_page.items():
        if page not in selected_pages and candidates:
            best = max(candidates, key=score_table_candidate)
            selected.append(best)

    selected.sort(key=lambda t: (t.get("page") or 9999, t.get("label") or ""))
    return selected


def table_dict_to_block(table: dict) -> str:
    lines = [
        "[TABLE_START]",
        f"Source: {table.get('source', 'unknown')}",
        f"Page: {table.get('page', 'unknown')}",
    ]
    label = table.get("label")
    if label:
        lines.append(f"Label: {label}")

    for row in table.get("rows", []):
        cleaned = [truncate_cell(cell) for cell in row]
        lines.append("\t".join(cleaned))

    lines.append("[TABLE_END]")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Marker integration (optional)
# ---------------------------------------------------------------------------

def find_marker_output(pdf_path: Path) -> Optional[Path]:
    """
    Look for a Marker-generated markdown file alongside the PDF.

    Marker typically outputs:  <stem>/<stem>.md  or  <stem>.marker.md
    We check several conventions.
    """
    stem = pdf_path.stem
    parent = pdf_path.parent

    candidates = [
        parent / f"{stem}.marker.md",
        parent / f"{stem}_marker.md",
        parent / stem / f"{stem}.md",
        parent / "marker_output" / f"{stem}.md",
        parent / "marker_output" / stem / f"{stem}.md",
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


def read_marker_output(marker_path: Path) -> str:
    """Read and lightly clean Marker markdown output."""
    text = marker_path.read_text(encoding="utf-8", errors="ignore")
    text = clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN)
    return text


# ---------------------------------------------------------------------------
# Document readers
# ---------------------------------------------------------------------------

def read_pdf(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    """
    Read a PDF. If a Marker output file exists alongside it, use that for
    the main text (better layout recovery) while still extracting tables
    via pdfplumber/camelot.
    """
    marker_path = find_marker_output(path)

    # Always extract tables with pdfplumber/camelot regardless of Marker
    plumber_tables = extract_tables_pdfplumber(path)
    camelot_tables = extract_tables_camelot(path)
    best_tables = deduplicate_and_select_best_tables(plumber_tables + camelot_tables)

    table_blocks = []
    for t in best_tables:
        p_num = t.get("page")
        block = table_dict_to_block(t)
        table_blocks.append((p_num, block))

    if marker_path:
        print(f"  Using Marker output: {marker_path}")
        marker_text = read_marker_output(marker_path)
        # Append table blocks that may not be in Marker output
        table_section = ""
        if table_blocks:
            table_parts = ["\n\n[Extracted tables (pdfplumber/camelot)]"]
            for _, block in table_blocks:
                table_parts.append(block)
            table_section = "\n".join(table_parts)
        return (marker_text + table_section).strip(), table_blocks

    # Fallback: pypdf text extraction with interleaved tables
    reader = PdfReader(str(path))
    table_map: Dict[int, List[str]] = {}
    for t in best_tables:
        p_num = t.get("page")
        block = table_dict_to_block(t)
        if p_num is not None:
            table_map.setdefault(p_num, []).append(block)

    pages = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as e:
            text = f"[Could not extract text from page {i}: {e}]"

        cleaned_page_text = clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN)
        page_parts = [f"\n\n[Page {i}]\n{cleaned_page_text}"]

        if i in table_map:
            page_parts.append("\n[Extracted tables]")
            page_parts.extend(table_map[i])

        pages.append("\n".join(part for part in page_parts if part.strip()))

    return "\n".join(pages).strip(), table_blocks


def read_txt(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    return clean_extracted_evidence_text(
        path.read_text(encoding="utf-8", errors="ignore"),
        domain=OUTPUT_DOMAIN,
    ), []


def read_docx(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    doc = Document(str(path))
    text = "\n".join(p.text for p in doc.paragraphs)
    return clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN), []


def read_csv_file(path: Path, max_rows: int = 120) -> Tuple[str, List[Tuple[int, str]]]:
    rows = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append("\t".join("" if x is None else str(x) for x in row))
    text = f"[CSV preview: first {min(len(rows), max_rows)} rows]\n" + "\n".join(rows)
    return clean_extracted_evidence_text(text, domain=OUTPUT_DOMAIN), []


def read_xlsx(path: Path, max_rows_per_sheet: int = 80) -> Tuple[str, List[Tuple[int, str]]]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"\n[Sheet: {ws.title}]")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                break
            out.append("\t".join("" if v is None else str(v) for v in row))
    return clean_extracted_evidence_text("\n".join(out), domain=OUTPUT_DOMAIN), []


def load_document(path: Path) -> Tuple[str, List[Tuple[int, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf(path)
    if suffix in {".txt", ".md"}:
        return read_txt(path)
    if suffix == ".docx":
        return read_docx(path)
    if suffix == ".csv":
        return read_csv_file(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path)
    raise ValueError(f"Unsupported file type: {path.name}")


def split_text(text: str, max_chars: int = MAX_SECTION_CHARS) -> List[str]:
    if not text:
        return ["[No extractable text found]"]
    if len(text) <= max_chars:
        return [text]
    paras = text.split("\n\n")
    chunks: List[str] = []
    current: List[str] = []
    current_len = 0
    for para in paras:
        para = para.strip()
        if not para:
            continue
        add_len = len(para) + 2
        if current and current_len + add_len > max_chars:
            chunks.append("\n\n".join(current).strip())
            current = [para]
            current_len = len(para)
        else:
            current.append(para)
            current_len += add_len
    if current:
        chunks.append("\n\n".join(current).strip())
    return chunks if chunks else ["[No extractable text found]"]


def collect_input_paths(paths: Iterable[str]) -> List[Path]:
    supported = {".pdf", ".txt", ".md", ".docx", ".csv", ".xlsx", ".xlsm"}
    results: List[Path] = []
    for raw in paths:
        p = Path(raw).expanduser()
        if p.is_dir():
            for child in sorted(p.rglob("*")):
                if child.is_file() and child.suffix.lower() in supported:
                    results.append(child)
        elif p.is_file() and p.suffix.lower() in supported:
            results.append(p)
    return results


# ---------------------------------------------------------------------------
# LLM interaction
# ---------------------------------------------------------------------------

def apply_qwen_chat_template(tokenizer, user_text: str) -> str:
    messages = [{"role": "user", "content": user_text}]
    return tokenizer.apply_chat_template(
        messages,
        tokenize=False,
        add_generation_prompt=True,
        enable_thinking=False,
    )


def make_default_sampler():
    return make_sampler(TEMPERATURE, top_p=TOP_P, top_k=TOP_K)


def clean_model_output(text: str) -> str:
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("```markdown", "").replace("```text", "").replace("```", "")
    text = re.sub(r" +", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


# ---------------------------------------------------------------------------
# System prompt (enhanced with method-awareness)
# ---------------------------------------------------------------------------

SYSTEM_STYLE = """You are a rigorous academic peer review assistant.

Core rules:
- Work strictly from the provided material.
- Distinguish clearly between:
  1. Directly supported by the text
  2. Reasonable inference
  3. Speculation / missing information
- Do not invent sample sizes, analyses, results, measures, design features, claims, contradictions, artefacts, or errors.
- Prioritise internal validity, measurement quality, causal assumptions, statistical reasoning, modelling choices, transparency, and reporting.
- Be critical but fair.
- Use concise markdown headings and bullet points.
- Do NOT use LaTeX or dollar-sign math notation.
- If tabular material is present, treat it as available evidence unless clearly unreadable.
- If a table is imperfectly extracted, say "partly extracted" or "some entries may be ambiguous", not "missing", unless it is genuinely absent.
- Ignore minor OCR artefacts in variable names (e.g., WHT.5R, WHT$5R, WHT^0.5R all refer to Waist/Height^0.5). Do not critique these as notation errors.
- Be aware that OCR sometimes extracts '=' as 'Z' (e.g., r Z 0.35 means r = 0.35). Interpret accordingly.
- EQUATION OCR RULE: Ignore garbled equations, subscripts, or superscripts (e.g., vconsi, u(3)0, LEV2bout). These are known PDF extraction failures. Do NOT claim equations are unreadable, missing, or unverifiable due to formatting.
- PEDANTRY RULE: Ignore trivial typographical inconsistencies in reference lists, citations, or software version numbers (e.g., MLwiN version 2.1 vs 2.10). Do not mention them in the review.
- PARADIGM RULE: MCMC is a computational method, not a paradigm. If MCMC, MLwiN, Gibbs sampling, or DIC are mentioned without explicit Bayesian terminology (e.g., priors, posteriors, Bayes factors), you MUST treat the analysis as STRICTLY FREQUENTIST using MCMC for estimation. Do not hedge, do not describe the framework as 'unclear', and do not critique the absence of Bayesian diagnostics.

Generality rule:
- Apply the same standards across frequentist, Bayesian, predictive, and causal analyses.
- Do not assume a method-specific problem unless the extracted material supports it.

Severity ladder:
- Prefer "could be clarified" over "is wrong" unless the text directly supports the stronger claim.
- Prefer "reported briefly" or "not fully detailed" over "unreported" when something is mentioned but not elaborated.
- Prefer "interpretation may need clarification" over "data entry artefact" unless there is direct evidence of corruption.
- Prefer "could not verify from the extracted material" over "missing" when extraction is partial.

Numeric provenance rule:
- Academic papers frequently cite statistics from OTHER studies for context (e.g., "In a previous study by X et al., N=405 decisions were analysed").
- Do NOT conflate numbers from cited/referenced studies with the current study's data.
- Do NOT claim a "discrepancy" between a number from a cited study and a number from the current study.
- Before identifying any numeric inconsistency, verify that both numbers refer to the SAME dataset/analysis within the current study.
- If a number appears near "et al.", a year in parentheses, or "previous study/larger dataset", it likely belongs to a different study.
"""

COMMON_DIAGNOSTIC_ALIASES = """Examples of reported diagnostics and checks across paradigms:
- frequentist/model-based: VIF, residual plots, heteroscedasticity checks, AIC, BIC, confidence intervals, sensitivity analyses, multiple imputation, cross-validation
- Bayesian: priors, prior predictive checks, posterior predictive checks, R-hat, effective sample size, divergences, LOO, WAIC, Bayes factors, prior sensitivity
- prediction/ML: calibration, discrimination, ROC/AUC, Brier score, internal validation, external validation, optimism correction
- causal/design: DAGs, assumptions, preregistration, protocol deviations, missing-data mechanisms
- MCMC as estimation (non-Bayesian): convergence diagnostics, burn-in, DIC, variance components with standard errors, Wald-type tests
- distributional/GAMLSS: worm plots, randomised quantile residuals, GAIC, distribution family comparison
"""


# ---------------------------------------------------------------------------
# Review prompts (method-aware)
# ---------------------------------------------------------------------------

def review_chunk(model, tokenizer, chunk: DocChunk, method_expectations: str = "",
                 manifest_summary: str = "") -> str:
    context_block = ""
    if method_expectations:
        context_block += f"\n{method_expectations}\n"
    if manifest_summary:
        context_block += f"\nEvidence manifest (what has been extracted from this file):\n{manifest_summary}\n"

    user_text = f"""
{SYSTEM_STYLE}

{COMMON_DIAGNOSTIC_ALIASES}
{context_block}
Review this manuscript chunk.

Source: {chunk.source_name}
Chunk: {chunk.chunk_id}

Return markdown under exactly these headings:
## Supported points
## Reasonable inferences
## Missing or unclear
## Statistical / methodological concerns
## Table-specific notes
## Questions raised by this chunk

Instructions:
- MCMC is an estimation method, not a paradigm. If the text mentions MCMC, MLwiN, or DIC but does not explicitly mention "priors" or "posteriors", it is a frequentist analysis. Do not evaluate it against Bayesian reporting standards and do not state that the framework is unclear.
- Where tables are present, use them as evidence.
- If a table appears partially extracted, say so explicitly.
- Do not claim a table, coefficient, equation, or diagnostic is missing if it is present in the chunk.
- Do not use LaTeX or dollar-sign notation.
- Be concise and avoid repetition.
- Do not ask for information already clearly reported in the chunk.
- Do not convert obvious PDF/OCR extraction artefacts into substantive methodological concerns.
- Do not critique garbled equations, subscripts, or superscripts (e.g., vconsi, u(3)0). Treat them as text extraction limits, not author errors.
- Do not complain about trivial typographical discrepancies in software versioning or references (e.g., MLwiN 2.1 vs 2.10).
- Do not describe a parameterisation choice or scaling convention as an error unless the text directly contradicts itself.
- Do not infer data entry artefacts from a small coefficient with a large test statistic alone.
- If an equation or model form is shown in the chunk, do not say the functional form is absent or unclear.
- Distinguish between "not reported", "reported briefly", and "reported but not fully interpretable from this chunk".
- If the text refers to an equation, figure, table, appendix item, or model specification that is not clearly recoverable from extraction, do not call it missing.
- Instead say that it appears to be present in the PDF but could not be evaluated confidently from the extracted material.
- Match your diagnostic expectations to the method classification provided above. Do not request diagnostics from the wrong framework.
- Do not request VIF for single-predictor models.

Material:
\"\"\"
{chunk.text}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=SECTION_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out)


def synthesize_file_review(model, tokenizer, file_name: str, combined_chunk_review: str,
                           method_expectations: str = "", manifest_summary: str = "") -> str:
    context_block = ""
    if method_expectations:
        context_block += f"\n{method_expectations}\n"
    if manifest_summary:
        context_block += f"\nEvidence manifest:\n{manifest_summary}\n"

    user_text = f"""
{SYSTEM_STYLE}

{COMMON_DIAGNOSTIC_ALIASES}
{context_block}
Create a concise file-level review summary from the chunk notes below.

File: {file_name}

Required headings:
# File synopsis
# Strongly supported strengths
# Strongly supported concerns
# Reporting limits or ambiguities
# Questions raised by this file

Rules:
- If a study uses MCMC or MLwiN without explicit Bayesian terminology (priors, posteriors), treat it as strictly frequentist. Do not suggest the paradigm is unclear and do not mention missing Bayesian elements.
- Do not include complaints about OCR-garbled equations or trivial software version discrepancies.
- Base the summary only on the chunk notes.
- Maximum 4 bullets under any heading.
- Prefer directly supported points over generic reviewer concerns.
- Do not say a table, coefficient, equation, or diagnostic is missing if the chunk notes contain it.
- If something is present but brief, say "reported briefly" or "not fully detailed", not "unreported".
- Do not use LaTeX or dollar-sign notation.
- Do not convert obvious PDF/OCR extraction artefacts into substantive concerns.
- Do not call a reporting choice a contradiction unless two directly conflicting statements are present.
- Do not infer data entry artefacts from small coefficients, exponent notation, or scaling choices alone.
- Finish cleanly after the final heading.
- Distinguish between "not present in the manuscript" and "not recoverably extracted".
- If an item is referenced in the manuscript text but not clearly accessible from extraction, describe it as "present but not fully accessible in the extracted material".
- Treat extraction failure as a review limitation, not automatically as a manuscript fault.
- Match diagnostic expectations to the method classification. Do not request diagnostics from the wrong paradigm.
- Consult the evidence manifest to avoid claiming something is absent when it is recorded as present.

Chunk notes:
\"\"\"
{combined_chunk_review}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=FILE_SYNTHESIS_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out)


def synthesize_report(model, tokenizer, file_summaries: List[Tuple[str, str]],
                      all_manifests: Optional[List[EvidenceManifest]] = None) -> str:
    joined = []
    for name, summary in file_summaries:
        joined.append(f"# File: {name}\n{summary}")
    joined_text = "\n\n".join(joined)

    manifest_block = ""
    if all_manifests:
        manifest_block = "\n\nEvidence manifests:\n" + "\n\n".join(m.summary_text() for m in all_manifests)

    user_text = f"""
{SYSTEM_STYLE}

{COMMON_DIAGNOSTIC_ALIASES}
{manifest_block}

Create one integrated peer-review report from the file summaries below.

Required headings:
# Overall synopsis
# Major strengths
# Major concerns
# Statistical and methodological issues
# Questions for the authors
# Suggested comments to editor
# Confidence and limits of this review

Rules:
- Treat studies using MCMC/MLwiN/DIC without explicit Bayesian terms as frequentist. Do not hedge, do not suggest the framework is unclear, and do not raise Bayesian-specific concerns.
- Do not include complaints about OCR-garbled equations or trivial software version discrepancies.
- Base the report only on the supplied file summaries.
- State clearly when a point is directly supported versus inferred.
- Do not overclaim.
- Maximum 4 bullets under any heading.
- Prefer concise bullets over long paragraphs.
- Do not repeat the same point across multiple headings unless necessary.
- Do not use LaTeX or dollar-sign notation.
- Do not claim that a table, coefficient, equation, or diagnostic is missing if it is present in the file summaries.
- Do not ask for a statistic as unreported if it is reported in the file summaries.
- If reporting is present but limited, say "reported briefly" or "reported without full detail".
- Do not convert obvious PDF/OCR extraction artefacts into substantive concerns.
- Do not call a reporting choice a contradiction unless two directly conflicting statements are present.
- Do not infer data entry artefacts from small coefficients, exponent notation, or scaling choices alone.
- If the model form or equation appears in the summaries, do not say the functional form is absent or unclear.
- Finish cleanly after the final heading.
- If a criticism depends on the absence of an equation, figure, table, appendix item, or model specification, include it only when the evidence summaries genuinely indicate absence.
- If the manuscript text points to such an item but extraction is inadequate, say it appears to be present in the PDF but could not be fully evaluated from the extracted material.
- Treat extraction limitations as limitations of this review unless the evidence supports a stronger claim.
- Match all diagnostic expectations to the classified method. Consult the evidence manifest.

File summaries:
\"\"\"
{joined_text}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=SYNTHESIS_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out)


def validate_report_against_evidence(
    model, tokenizer, report_text: str,
    file_summaries: List[Tuple[str, str]],
    programmatic_corrections: Optional[List[str]] = None,
) -> str:
    joined = []
    for name, summary in file_summaries:
        joined.append(f"# File: {name}\n{summary}")
    evidence_text = "\n\n".join(joined)

    correction_block = ""
    if programmatic_corrections:
        correction_block = (
            "\n\nPROGRAMMATIC CORRECTIONS (these have been verified automatically and MUST be applied):\n"
            + "\n".join(f"- {c}" for c in programmatic_corrections)
            + "\n"
        )

    user_text = f"""You are checking a peer-review report for factual consistency and proportionality against extracted evidence summaries.

Task:
Revise the report only where it incorrectly claims that information is missing, absent, contradictory, erroneous, or unreported.

General rule:
- Do not criticise the absence of something that is actually reported.
- Only criticise adequacy, transparency, or interpretation when the evidence supports that criticism.
- Actively REMOVE any critique regarding garbled equations, unreadable mathematical specifications, or OCR artifacts in formulas.
- Actively REMOVE any critique regarding trivial software version discrepancies (e.g., 2.1 vs 2.10) or reference formatting.

{COMMON_DIAGNOSTIC_ALIASES}
{correction_block}
Important rules:
- If the evidence mentions MCMC, Gibbs sampling, MLwiN, or DIC without explicit Bayesian terms, you MUST treat it as a frequentist model. Actively remove any sentences that hedge (e.g., "it is unclear if it is Bayesian") or that critique missing Bayesian diagnostics.
- Treat both extracted table blocks and plain chunk text containing tabular rows as evidence.
- If tables, coefficients, standard errors, confidence intervals, fit statistics, equations, or diagnostics are present in the evidence summaries, do not describe them as missing, absent, or unverifiable.
- If something is present but concise, revise wording to "reported briefly", "not fully detailed", or "could not be fully evaluated from the extracted material".
- If a model form appears in the evidence summaries, do not say the functional form is undefined.
- Do not convert obvious PDF/OCR extraction artefacts into substantive methodological concerns.
- Do not infer data entry artefacts from small coefficients, exponent notation, or scaling choices alone.
- Do not describe table numbering as inconsistent unless the evidence summaries themselves contain conflicting labels.
- Do not treat notation or terminology conventions as a major concern unless they create genuine inferential ambiguity.
- Replace overly strong wording with proportionate wording when needed.
- Preserve valid critical points.
- Do not introduce new criticisms.
- Do not use LaTeX or dollar-sign notation.
- Return the revised report only.
- If the evidence summaries refer to Eq., Figure, Table, Appendix, Supplementary Table, or similar labels, do not say the item is missing unless the evidence explicitly indicates absence.
- Revise such claims to wording like "the item appears to be present in the PDF but could not be fully evaluated from the extracted material".
- Treat extraction failure as a review limitation rather than evidence that the manuscript omitted the item.
- Where the manuscript clearly points to an equation, table, figure, or appendix item that was not recoverably extracted, treat this as a limitation of the extracted material rather than evidence that the item is absent.
- Apply all PROGRAMMATIC CORRECTIONS listed above. These are verified and must be implemented.

Report:
\"\"\"
{report_text}
\"\"\"

Evidence summaries:
\"\"\"
{evidence_text}
\"\"\"
""".strip()
    prompt = apply_qwen_chat_template(tokenizer, user_text)
    sampler = make_default_sampler()
    out = generate(
        model, tokenizer, prompt=prompt,
        max_tokens=VALIDATION_MAX_TOKENS, sampler=sampler, verbose=False
    )
    return clean_model_output(out)



def correct_review_contradictions(
    report_text: str,
    combined_text: str,
    table_blocks: list[str],
) -> str:
    """
    Light post-processing pass to soften review claims that directly
    contradict extracted evidence.
    """
    text_low = combined_text.lower()
    tables_low = "\n\n".join(table_blocks).lower()
    full_text = text_low + "\n\n" + tables_low
    revised = report_text

    # Confidence intervals present in tables/text
    if any(k in full_text for k in [
        "95% ci", "confidence interval", "lower bound", "upper bound", " lower ", " upper "
    ]):
        revised = revised.replace(
            "confidence intervals are absent from the extracted text",
            "confidence interval information appears in the extracted material, although fuller reporting may still be desirable",
        )
        revised = revised.replace(
            "does not contain confidence intervals",
            "contains some confidence interval information",
        )
        revised = revised.replace(
            "lacks confidence intervals",
            "does not always present confidence intervals as prominently as it could",
        )
    # Strong table visibility rule:
    # if structured table evidence contains estimates, bounds, and inferential columns,
    # do not allow the review to claim that the main numerical values are not visible.
    table_has_main_numeric_columns = (
        "estimate" in full_text
        and "lower" in full_text
        and "upper" in full_text
        and ("effect size" in full_text or "\tp\t" in full_text or " p " in full_text)
    )

    if table_has_main_numeric_columns:
        replacements = [
            (
                "specific numerical values for confidence intervals and individual data points are not fully recoverable from the extracted text chunks",
                "substantial numerical values, including interval bounds and effect estimates, are recoverable from the extracted tables",
            ),
            (
                "specific numerical values for trimmed means and confidence interval bounds are not visible in the extracted text chunks",
                "trimmed means and confidence interval bounds are visible in the extracted tables",
            ),
            (
                "the actual numerical data are not fully extracted in the provided text",
                "substantial numerical table data are present in the extracted material",
            ),
            (
                "specific numerical values for confidence intervals are not visible in the extracted text",
                "confidence interval values are visible in the extracted tables",
            ),
            (
                "specific numerical values are not visible in the extracted text",
                "specific numerical values are visible in the extracted tables",
            ),
            (
                "not fully recoverable from the extracted text chunks",
                "recoverable from the extracted tables",
            ),
            (
                "not visible in the extracted text chunks",
                "visible in the extracted tables",
            ),
            (
                "preventing a direct verification of the reported summary statistics",
                "allowing at least partial direct verification of the reported summary statistics",
            ),
            (
                "limiting the ability to verify the precision of the reported results",
                "still allowing substantial verification of the precision of the reported results from the extracted tables",
            ),
        ]
        for old, new in replacements:
            revised = revised.replace(old, new)

        # Soften broader claims if they still survive
        revised = revised.replace(
            "the extracted text does not show the numerical results",
            "the extracted tables show substantial numerical results",
        )
        revised = revised.replace(
            "the extracted material does not show the numerical results",
            "the extracted material includes substantial numerical results in the tables",
        )

    # Model specification / equations present
    if any(k in full_text for k in [
        "the following equation",
        "equation was applied",
        "model equation",
        "ln(vo2max",
        "p = w",
        "work-time model",
        "work–time model",
    ]):
        revised = revised.replace(
            "specific model specifications and detailed diagnostics are not fully detailed in the extracted text",
            "the extracted text contains at least a partial model specification, though some details may still be brief",
        )
        revised = revised.replace(
            "model specification is absent",
            "model specification is at least partly present",
        )

    # Standard errors / SD-like reporting present
    if any(k in full_text for k in [
        "\tse\t", " se ", "standard error", "trimmed sd", "sd (0.2)"
    ]):
        revised = revised.replace(
            "standard errors are absent",
            "standard errors are at least partly present in the extracted material",
        )
        revised = revised.replace(
            "the extracted material does not contain standard errors",
            "the extracted material contains at least some standard error or SD-like information",
        )

    # Structured table evidence present: estimates, bounds, p-values, effect sizes
    table_has_ci_like = all(k in full_text for k in ["estimate", "lower", "upper"])
    table_has_inference_like = ("p" in full_text and "effect size" in full_text)

    if table_has_ci_like or table_has_inference_like:
        replacements = [
            (
                "the actual numerical data are not fully extracted in the provided text",
                "substantial numerical table data are present in the extracted material",
            ),
            (
                "preventing a direct verification of the reported summary statistics",
                "allowing at least partial direct verification of the reported summary statistics",
            ),
            (
                "the tables are not accessible",
                "the tables are at least partly accessible in the extracted material",
            ),
            (
                "does not contain confidence intervals",
                "contains confidence interval information in the extracted tables",
            ),
            (
                "confidence intervals are absent from the extracted text",
                "confidence interval information appears in the extracted tables and text",
            ),
            (
                "the extracted material does not contain standard errors",
                "the extracted material contains at least some uncertainty information in the tables",
            ),
        ]
        for old, new in replacements:
            revised = revised.replace(old, new)

    return revised

# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_report(output_dir: Path, input_paths: List[Path], report_text: str,
                 manifests: Optional[List[EvidenceManifest]] = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    stem = input_paths[0].stem if len(input_paths) == 1 else "multi_file_review"
    report_path = output_dir / f"{stem}_review_{stamp}.md"
    sources = "\n".join(f"- {p}" for p in input_paths)

    manifest_section = ""
    if manifests:
        manifest_lines = []
        for m in manifests:
            mc = m.method_class.value if m.method_class else "unclassified"
            manifest_lines.append(f"- **{m.source_name}**: method={mc}, tables={m.n_tables}, "
                                  f"model_spec={m.has_model_spec}, equations={m.has_equations}, "
                                  f"SEs={m.has_standard_errors}, variance_components={m.has_variance_components}, "
                                  f"fit_stats={m.has_model_fit_stats}")
        manifest_section = "\nEvidence summary:\n" + "\n".join(manifest_lines) + "\n"

    header = f"""# Local peer-review report

Generated: {datetime.now().isoformat(timespec="seconds")}
Model: {MODEL_NAME}

Input files:
{sources}
{manifest_section}
---

"""
    report_path.write_text(header + report_text, encoding="utf-8")
    return report_path


def write_evidence_appendix(
    output_dir: Path,
    input_paths: List[Path],
    per_file_tables: Dict[str, List[Tuple[int, str]]],
    per_file_chunks: Dict[str, List[str]],
    manifests: Optional[List[EvidenceManifest]] = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    stem = input_paths[0].stem if len(input_paths) == 1 else "multi_file_review"
    appendix_path = output_dir / f"{stem}_evidence_appendix_{stamp}.md"

    lines: List[str] = []
    lines.append("# Evidence appendix")
    lines.append("")
    lines.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"Model: {MODEL_NAME}")
    lines.append("")
    lines.append("## Input files")
    for p in input_paths:
        lines.append(f"- {p}")
    lines.append("")

    # Manifest summary section
    if manifests:
        lines.append("## Evidence manifests")
        lines.append("")
        for m in manifests:
            lines.append(f"### {m.source_name}")
            lines.append("")
            lines.append("```text")
            lines.append(m.summary_text())
            lines.append("```")
            lines.append("")
            # Block type breakdown
            type_groups: Dict[str, List[str]] = {}
            for b in m.blocks:
                type_groups.setdefault(b.block_type.name, []).append(
                    b.text[:120].replace("\n", " ") + ("..." if len(b.text) > 120 else "")
                )
            for btype, previews in type_groups.items():
                lines.append(f"**{btype}** ({len(previews)} blocks)")
                lines.append("")
                for prev in previews[:5]:  # Show up to 5 previews per type
                    lines.append(f"- {prev}")
                if len(previews) > 5:
                    lines.append(f"- ... and {len(previews) - 5} more")
                lines.append("")

    for p in input_paths:
        name = p.name
        lines.append(f"# File: {name}")
        lines.append("")

        tables = per_file_tables.get(name, [])
        if tables:
            lines.append("## Extracted tables")
            lines.append("")
            for page_num, block in tables:
                lines.append(f"### Page {page_num}")
                lines.append("")
                lines.append("```text\n" + block + "\n```")
                lines.append("")
        else:
            lines.append("## Extracted tables")
            lines.append("")
            marker_path = find_marker_output(p) if p.suffix.lower() == ".pdf" else None
            if marker_path:
                lines.append("No extracted table blocks recorded (Visual parser may have embedded them in chunks).")
            else:
                lines.append("No extracted table blocks recorded.")
            lines.append("")

        chunks = per_file_chunks.get(name, [])
        if chunks:
            lines.append("## Chunk previews")
            lines.append("")
            for i, chunk_text in enumerate(chunks, start=1):
                preview = chunk_text[:MAX_APPENDIX_CHUNK_PREVIEW]
                if len(chunk_text) > MAX_APPENDIX_CHUNK_PREVIEW:
                    preview += "\n...[truncated preview]..."
                lines.append(f"### Chunk {i}")
                lines.append("")
                lines.append("```text\n" + preview + "\n```")
                lines.append("")
        else:
            lines.append("## Chunk previews")
            lines.append("")
            lines.append("No chunk previews available.")
            lines.append("")

    appendix_path.write_text("\n".join(lines), encoding="utf-8")
    return appendix_path
    
def correct_direction_of_effect_summaries(
    report_text: str,
    combined_text: str,
    table_blocks: list[str],
) -> str:
    """
    Conservative post-processing pass to reduce obvious direction-of-effect
    reversals in summaries when extracted evidence strongly supports a direction.

    This function is intentionally heuristic and only acts when:
    1. paired condition labels are clearly present, and
    2. the extracted text contains directional wording or table-style evidence.
    """
    full_text = (combined_text + "\n\n" + "\n\n".join(table_blocks)).lower()
    revised = report_text

    def has_any(*phrases: str) -> bool:
        return any(p.lower() in full_text for p in phrases)

    # Common paired labels seen in manuscripts
    paired_labels = [
        ("indoor", "outdoor"),
        ("control", "intervention"),
        ("control", "treatment"),
        ("placebo", "treatment"),
        ("placebo", "intervention"),
        ("male", "female"),
        ("pre", "post"),
        ("before", "after"),
        ("baseline", "follow-up"),
        ("baseline", "post"),
        ("rest", "exercise"),
        ("condition a", "condition b"),
    ]

    # Strong general “no difference” cues
    evidence_no_difference = has_any(
        "no difference was observed",
        "no significant difference was observed",
        "a reliable difference could not be established",
        "did not differ",
        "no effect was observed",
        "nonsignificant relative differences",
        "no statistically significant effect",
    )

    # General table-like structure cues
    has_estimate_structure = all(
        key in full_text for key in ["estimate", "lower", "upper"]
    )

    # Build generic directional evidence from paired labels
    for left, right in paired_labels:
        left_in_text = left in full_text
        right_in_text = right in full_text
        if not (left_in_text and right_in_text):
            continue

        evidence_right_higher = has_any(
            f"higher {right} compared with {left}",
            f"greater {right} compared with {left}",
            f"{right} was higher than {left}",
            f"{right} were higher than {left}",
            f"{right} condition was higher",
            f"{right} condition were higher",
        )

        evidence_left_higher = has_any(
            f"higher {left} compared with {right}",
            f"greater {left} compared with {right}",
            f"{left} was higher than {right}",
            f"{left} were higher than {right}",
            f"{left} condition was higher",
            f"{left} condition were higher",
        )

        # If table-style evidence exists and text explicitly states a direction,
        # do not allow the review to assert the opposite direction.
        if has_estimate_structure:
            if evidence_right_higher:
                revised = revised.replace(
                    f"higher {left} compared with {right}",
                    f"higher {right} compared with {left}",
                )
                revised = revised.replace(
                    f"greater {left} compared with {right}",
                    f"greater {right} compared with {left}",
                )
                revised = revised.replace(
                    f"{left} was higher than {right}",
                    f"{right} was higher than {left}",
                )
                revised = revised.replace(
                    f"{left} were higher than {right}",
                    f"{right} were higher than {left}",
                )

            if evidence_left_higher:
                revised = revised.replace(
                    f"higher {right} compared with {left}",
                    f"higher {left} compared with {right}",
                )
                revised = revised.replace(
                    f"greater {right} compared with {left}",
                    f"greater {left} compared with {right}",
                )
                revised = revised.replace(
                    f"{right} was higher than {left}",
                    f"{left} was higher than {right}",
                )
                revised = revised.replace(
                    f"{right} were higher than {left}",
                    f"{left} were higher than {right}",
                )

    # Soften over-strong summary language where evidence explicitly says no difference
    if evidence_no_difference:
        soft_replacements = [
            (
                "was significantly different",
                "was reported as different only where supported by the extracted material",
            ),
            (
                "showed a clear difference",
                "showed a difference only in some outcomes",
            ),
            (
                "demonstrated a clear effect",
                "demonstrated an effect only where supported by the extracted material",
            ),
            (
                "there was a clear difference",
                "there was a difference in some outcomes, while others showed no clear difference",
            ),
        ]
        for old, new in soft_replacements:
            revised = revised.replace(old, new)
            revised = revised.replace("in the extracted material in the extracted text", "in the extracted material")
            revised = revised.replace("in the extracted text in the extracted material", "in the extracted material")

    return revised

def enforce_negative_constraints(report_text: str) -> str:
    """
    Final light-touch cleanup pass.

    Purpose:
    - soften absolute absence claims that are often too strong
    - avoid overclaiming based on extraction limitations
    - keep wording conservative rather than rewriting substance
    """
    revised = report_text

    replacements = [
        (
            "is absent from the extracted text",
            "is not clearly identifiable in the extracted text",
        ),
        (
            "are absent from the extracted text",
            "are not clearly identifiable in the extracted text",
        ),
        (
            "is absent from the extracted material",
            "is not clearly identifiable in the extracted material",
        ),
        (
            "are absent from the extracted material",
            "are not clearly identifiable in the extracted material",
        ),
        (
            "is missing from the extracted text",
            "is not clearly recovered in the extracted text",
        ),
        (
            "are missing from the extracted text",
            "are not clearly recovered in the extracted text",
        ),
        (
            "is missing from the extracted material",
            "is not clearly recovered in the extracted material",
        ),
        (
            "are missing from the extracted material",
            "are not clearly recovered in the extracted material",
        ),
        (
            "not reported",
            "not clearly reported in the extracted material",
        ),
        (
            "not provided",
            "not clearly provided in the extracted material",
        ),
        (
            "cannot be verified",
            "cannot be fully verified from the extracted material",
        ),
        (
            "preventing direct verification",
            "limiting direct verification",
        ),
    ]

    for old, new in replacements:
        revised = revised.replace(old, new)

    revised = revised.replace(
        "the manuscript does not contain",
        "the extracted material does not clearly contain",
    )
    revised = revised.replace(
        "the manuscript lacks",
        "the extracted material does not clearly show",
    )
    revised = revised.replace(
        "there is no",
        "there is no clearly recovered",
    )

    return revised

# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description="Local manuscript review pipeline using MLX + Qwen3.5")
    parser.add_argument("inputs", nargs="+", help="One or more files or folders")
    parser.add_argument(
        "--output-dir",
        default=str(Path.home() / "local-llm/qwen35-review/reports"),
        help="Directory for markdown reports",
    )
    parser.add_argument(
        "--domain",
        default="general",
        help="Domain hint for OCR cleaning (e.g., exercise_physiology)",
    )
    args = parser.parse_args()

    global OUTPUT_DOMAIN
    OUTPUT_DOMAIN = args.domain

    input_paths = collect_input_paths(args.inputs)
    if not input_paths:
        print("No supported input files found.", file=sys.stderr)
        return 1

    print("Loading model...")
    model, tokenizer = load(MODEL_NAME)

    file_summaries: List[Tuple[str, str]] = []
    per_file_tables: Dict[str, List[Tuple[int, str]]] = {}
    per_file_chunks: Dict[str, List[str]] = {}
    all_manifests: List[EvidenceManifest] = []

    for path in input_paths:
        print(f"Reading: {path}")
        try:
            text, table_blocks = load_document(path)
        except Exception as e:
            file_summaries.append((path.name, f"Failed to read file: {e}"))
            per_file_tables[path.name] = []
            per_file_chunks[path.name] = []
            continue

        per_file_tables[path.name] = table_blocks

        # --- Phase 2: Structure evidence ---
        print(f"  Structuring evidence for {path.name}...")
        manifest = structure_evidence(path.name, text, table_blocks)
        all_manifests.append(manifest)

        mc = manifest.method_class
        print(f"  Method classification: {mc.value}")
        print(f"  Evidence: {manifest.n_tables} tables, model_spec={manifest.has_model_spec}, "
              f"equations={manifest.has_equations}, p_values={manifest.has_p_values}")

        # --- Phase 3: Get method-specific expectations ---
        method_expectations = get_method_expectations(mc)
        manifest_summary = manifest.summary_text()

        chunks = split_text(text)
        per_file_chunks[path.name] = chunks

        if not chunks:
            file_summaries.append((path.name, "No usable text extracted from this file."))
            continue

        # --- Phase 4: Method-aware chunk review ---
        chunk_outputs = []
        for i, chunk_text in enumerate(chunks, start=1):
            print(f"  Reviewing {path.name} chunk {i}/{len(chunks)}")
            chunk = DocChunk(source_name=path.name, chunk_id=i, text=chunk_text)
            try:
                reviewed = review_chunk(
                    model, tokenizer, chunk,
                    method_expectations=method_expectations,
                    manifest_summary=manifest_summary,
                )
            except Exception as e:
                reviewed = f"Chunk review failed: {e}"
            chunk_outputs.append(reviewed)

        combined = "\n\n".join(f"### Chunk {i}\n{txt}" for i, txt in enumerate(chunk_outputs, start=1))

        # --- Phase 5: File-level synthesis ---
        print(f"  Synthesising file-level review for {path.name}...")
        try:
            file_summary = synthesize_file_review(
                model, tokenizer, path.name, combined,
                method_expectations=method_expectations,
                manifest_summary=manifest_summary,
            )
        except Exception as e:
            file_summary = f"# File synopsis\nFile-level synthesis failed: {e}"
        file_summaries.append((path.name, file_summary))

    # --- Phase 5 continued: Final synthesis ---
    print("Synthesising final report...")
    try:
        final_report = synthesize_report(model, tokenizer, file_summaries, all_manifests=all_manifests)

        # --- Phase 6: Programmatic post-checks ---
        all_corrections: List[str] = []
        for manifest in all_manifests:
            corrections = programmatic_post_checks(final_report, manifest)
            all_corrections.extend(corrections)

        if all_corrections:
            print(f"  Programmatic checks found {len(all_corrections)} correction(s):")
            for c in all_corrections:
                print(f"    - {c[:100]}...")

        # --- Phase 6 continued: LLM validation ---
        print("Validating final report against evidence...")
        final_report = validate_report_against_evidence(
            model,
            tokenizer,
            final_report,
            file_summaries,
            programmatic_corrections=all_corrections if all_corrections else None,
        )

        # --- Phase 6b: Light contradiction correction ---
        all_text = "\n\n".join(text for _, text in file_summaries)
        all_tables = [tbl for tables in per_file_tables.values() for _, tbl in tables]

        final_report = correct_review_contradictions(
            final_report,
            all_text,
            all_tables,
        )

        # --- Phase 6c: Direction-of-effect consistency guard ---
        final_report = correct_direction_of_effect_summaries(
            final_report,
            all_text,
            all_tables,
        )

        # --- Phase 6d: Conservative wording cleanup ---
        final_report = enforce_negative_constraints(final_report)

        # --- Phase 6e: Remove leaked markdown/math artifacts ---
        final_report = clean_markdown_math_artifacts(final_report)

    except Exception as e:
        final_report = (
            "# Overall synopsis\n"
            f"Synthesis failed: {e}\n\n"
            "# File-level notes\n\n"
            + "\n\n".join(f"## {name}\n{summary}" for name, summary in file_summaries)
        )
    output_dir = Path(args.output_dir)
    report_path = write_report(output_dir, input_paths, final_report, manifests=all_manifests)
    appendix_path = write_evidence_appendix(
        output_dir, input_paths, per_file_tables, per_file_chunks, manifests=all_manifests,
    )

    print(f"Saved report: {report_path}")
    print(f"Saved evidence appendix: {appendix_path}")
    return 0

def clean_markdown_math_artifacts(report_text: str) -> str:
    """
    Remove light LaTeX-style math delimiters that leak into Markdown output.
    Keeps the text readable without trying to do full LaTeX parsing.
    """
    revised = report_text

    replacements = {
        "$W'$": "W′",
        "$W’$": "W′",
        "$W′$": "W′",
        "$R^2$": "R²",
        "$r^2$": "r²",
        "$p$": "p",
        "$P$": "P",
        "$CP$": "CP",
        "$VO2max$": "VO2max",
        "$VO_2$": "VO2",
        "$VO_{2max}$": "VO2max",
    }

    for old, new in replacements.items():
        revised = revised.replace(old, new)

    # remove remaining inline math delimiters conservatively
    revised = re.sub(r"\$([^$\n]{1,80})\$", r"\1", revised)

    # common exponent cleanups
    revised = revised.replace("R^2", "R²")
    revised = revised.replace("r^2", "r²")

    # light cleanup of common inequalities
    revised = revised.replace("$<$", "<")
    revised = revised.replace("$>$", ">")
    revised = revised.replace("$<0$", "<0")
    revised = revised.replace("$>0$", ">0")

    return revised
    
if __name__ == "__main__":
    raise SystemExit(main())

